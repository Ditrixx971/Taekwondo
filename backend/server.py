from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import random
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

# ============ MODELS ============

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "coach"  # coach, admin, or master

class UserLogin(BaseModel):
    email: str
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    role: str
    picture: Optional[str] = None

# ============ COMPETITION MODELS ============

class Competition(BaseModel):
    model_config = ConfigDict(extra="ignore")
    competition_id: str = Field(default_factory=lambda: f"comp_{uuid.uuid4().hex[:12]}")
    nom: str
    date: str  # YYYY-MM-DD
    lieu: str
    heure_debut: str = "09:00"
    duree_estimee_heures: int = 8
    statut: str = "active"  # active, terminee, annulee
    coaches_autorises: List[str] = []  # Liste des user_id des coachs autorisés
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str = ""

class CompetitionCreate(BaseModel):
    nom: str
    date: str
    lieu: str
    heure_debut: str = "09:00"
    duree_estimee_heures: int = 8
    coaches_autorises: List[str] = []

# ============ COMPETITEUR MODELS ============

class Competiteur(BaseModel):
    model_config = ConfigDict(extra="ignore")
    competiteur_id: str = Field(default_factory=lambda: f"cptr_{uuid.uuid4().hex[:12]}")
    competition_id: str  # Lié à une compétition
    nom: str
    prenom: str
    date_naissance: str
    sexe: str  # M or F
    poids_declare: float  # Poids déclaré par le coach à l'inscription
    poids_officiel: Optional[float] = None  # Poids officiel après pesée
    club: str
    categorie_id: Optional[str] = None
    surclasse: bool = False  # Si le compétiteur est surclassé dans une catégorie supérieure
    pese: bool = False  # Statut de pesée
    disqualifie: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str = ""

class CompetiteurCreate(BaseModel):
    competition_id: str
    nom: str
    prenom: str
    date_naissance: str
    sexe: str
    poids_declare: float
    club: str
    surclasse: bool = False  # Si le compétiteur est surclassé
    categorie_surclasse_id: Optional[str] = None  # Catégorie choisie si surclassé

class PeseeUpdate(BaseModel):
    poids_officiel: float

class Categorie(BaseModel):
    model_config = ConfigDict(extra="ignore")
    categorie_id: str = Field(default_factory=lambda: f"cat_{uuid.uuid4().hex[:12]}")
    competition_id: str  # Lié à une compétition
    nom: str
    age_min: int
    age_max: int
    sexe: str
    poids_min: float
    poids_max: float

class CategorieCreate(BaseModel):
    competition_id: str
    nom: str
    age_min: int
    age_max: int
    sexe: str
    poids_min: float
    poids_max: float

class AireCombat(BaseModel):
    model_config = ConfigDict(extra="ignore")
    aire_id: str = Field(default_factory=lambda: f"aire_{uuid.uuid4().hex[:12]}")
    competition_id: str  # Lié à une compétition
    nom: str
    numero: int
    statut: str = "active"  # active, pause, hs (hors service)

class AireCombatCreate(BaseModel):
    competition_id: str
    nom: str
    numero: int

class AireCombatUpdate(BaseModel):
    nom: Optional[str] = None
    statut: Optional[str] = None  # active, pause, hs

# Garder Tatami pour rétrocompatibilité (alias)
class Tatami(BaseModel):
    model_config = ConfigDict(extra="ignore")
    tatami_id: str = Field(default_factory=lambda: f"tat_{uuid.uuid4().hex[:12]}")
    competition_id: str
    nom: str
    numero: int

class TatamiCreate(BaseModel):
    competition_id: str
    nom: str
    numero: int

class Combat(BaseModel):
    model_config = ConfigDict(extra="ignore")
    combat_id: str = Field(default_factory=lambda: f"cbt_{uuid.uuid4().hex[:12]}")
    competition_id: str  # Lié à une compétition
    categorie_id: str
    aire_id: Optional[str] = None  # Aire de combat assignée
    tatami_id: Optional[str] = None  # Alias rétrocompatibilité
    tour: str  # seizieme, huitieme, quart, demi, finale (PAS de bronze en taekwondo)
    position: int  # position in bracket
    has_bye: bool = False  # Si ce combat contient un BYE (un seul combattant)
    ordre: int = 0  # ordre d'exécution global
    rouge_id: Optional[str] = None
    bleu_id: Optional[str] = None
    vainqueur_id: Optional[str] = None
    score_rouge: int = 0
    score_bleu: int = 0
    type_victoire: Optional[str] = None  # normal, forfait, abandon, disqualification, non_dispute
    statut: str = "a_venir"  # a_venir, en_cours, termine, non_dispute
    termine: bool = False
    est_finale: bool = False  # Si c'est un combat de finale (à faire en dernier)
    heure_debut: Optional[str] = None  # ISO format
    duree_minutes: int = 6  # durée estimée en minutes
    est_pause: bool = False  # si c'est un créneau de pause
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # ============ CHAMPS POUR COMBATS MANUELS ============
    mode_creation: str = "auto"  # auto (tirage au sort) ou manuel
    combat_suivant_id: Optional[str] = None  # Combat où le vainqueur doit aller
    combat_suivant_slot: Optional[str] = None  # "rouge" ou "bleu" - position dans le combat suivant
    combat_source_rouge_id: Optional[str] = None  # Combat dont le vainqueur devient rouge
    combat_source_bleu_id: Optional[str] = None  # Combat dont le vainqueur devient bleu  
    pret: bool = False  # Si le combat est prêt (deux combattants définis et disponibles)
    nom_personnalise: Optional[str] = None  # Nom personnalisé pour le combat (ex: "Finale consolante")

class CombatResultat(BaseModel):
    vainqueur_id: str
    score_rouge: int = 0
    score_bleu: int = 0
    type_victoire: str = "normal"

# ============ MODÈLES POUR COMBATS MANUELS ============

class CombatManuelCreate(BaseModel):
    """Création d'un combat manuel"""
    competition_id: str
    categorie_id: str
    tour: str = "manuel"  # Type de tour ou "manuel"
    position: int = 1
    rouge_id: Optional[str] = None  # Compétiteur rouge (optionnel si connexion)
    bleu_id: Optional[str] = None  # Compétiteur bleu (optionnel si connexion)
    nom_personnalise: Optional[str] = None  # Nom personnalisé
    aire_id: Optional[str] = None  # Aire assignée

class CombatManuelUpdate(BaseModel):
    """Mise à jour d'un combat manuel"""
    tour: Optional[str] = None
    position: Optional[int] = None
    rouge_id: Optional[str] = None
    bleu_id: Optional[str] = None
    nom_personnalise: Optional[str] = None
    aire_id: Optional[str] = None
    ordre: Optional[int] = None

class ConnexionCombatRequest(BaseModel):
    """Connecter le vainqueur d'un combat vers un autre"""
    combat_source_id: str  # Combat dont le vainqueur sera envoyé
    combat_destination_id: str  # Combat où le vainqueur ira
    slot_destination: str  # "rouge" ou "bleu"

class DeconnexionCombatRequest(BaseModel):
    """Déconnecter un combat source d'un slot"""
    combat_destination_id: str
    slot: str  # "rouge" ou "bleu"

class AssignerAireRequest(BaseModel):
    """Assigner un combat à une aire"""
    combat_id: str
    aire_id: str

class PlanificationCreate(BaseModel):
    heure_debut_competition: str  # ISO format ex: "09:00"
    duree_combat_minutes: int = 6
    pauses: List[dict] = []  # [{"apres_combat": 10, "duree_minutes": 15}]

class ModifierOrdreCombat(BaseModel):
    combat_id: str
    nouvel_ordre: int
    nouvelle_heure: Optional[str] = None

class Medaille(BaseModel):
    model_config = ConfigDict(extra="ignore")
    medaille_id: str = Field(default_factory=lambda: f"med_{uuid.uuid4().hex[:12]}")
    categorie_id: str
    competiteur_id: str
    type: str  # or, argent, bronze

class HistoriqueResultat(BaseModel):
    model_config = ConfigDict(extra="ignore")
    historique_id: str = Field(default_factory=lambda: f"hist_{uuid.uuid4().hex[:12]}")
    combat_id: str
    ancien_vainqueur_id: Optional[str]
    nouveau_vainqueur_id: str
    ancien_score_rouge: int
    ancien_score_bleu: int
    nouveau_score_rouge: int
    nouveau_score_bleu: int
    modifie_par: str
    modifie_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============ AUTH HELPERS ============

async def get_current_user(request: Request) -> User:
    session_token = request.cookies.get("session_token")
    auth_header = request.headers.get("Authorization")
    
    if auth_header and auth_header.startswith("Bearer "):
        session_token = auth_header.split(" ")[1]
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Non authentifié")
    
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Session invalide")
    
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expirée")
    
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
    
    return User(**user)

async def require_admin(user: User = Depends(get_current_user)) -> User:
    if user.role not in ["admin", "master"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    return user

async def require_master(user: User = Depends(get_current_user)) -> User:
    if user.role != "master":
        raise HTTPException(status_code=403, detail="Accès réservé au super-administrateur (MASTER)")
    return user

# ============ AUTH ENDPOINTS ============

@api_router.post("/auth/register")
async def register(data: UserCreate, response: Response):
    existing = await db.users.find_one({"email": data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    from passlib.hash import bcrypt
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    user_doc = {
        "user_id": user_id,
        "email": data.email,
        "password": bcrypt.hash(data.password),
        "name": data.name,
        "role": data.role,
        "picture": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    session_token = f"sess_{uuid.uuid4().hex}"
    session_doc = {
        "session_id": f"sid_{uuid.uuid4().hex[:12]}",
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session_doc)
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    return {"user_id": user_id, "email": data.email, "name": data.name, "role": data.role}

@api_router.post("/auth/login")
async def login(data: UserLogin, response: Response):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    from passlib.hash import bcrypt
    if not bcrypt.verify(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    session_token = f"sess_{uuid.uuid4().hex}"
    session_doc = {
        "session_id": f"sid_{uuid.uuid4().hex[:12]}",
        "user_id": user["user_id"],
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session_doc)
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    return {
        "user_id": user["user_id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "picture": user.get("picture")
    }

@api_router.post("/auth/session")
async def exchange_session(request: Request, response: Response):
    body = await request.json()
    session_id = body.get("session_id")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id requis")
    
    async with httpx.AsyncClient() as client_http:
        resp = await client_http.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id}
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Session invalide")
        oauth_data = resp.json()
    
    email = oauth_data["email"]
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    
    if existing:
        user_id = existing["user_id"]
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"name": oauth_data["name"], "picture": oauth_data.get("picture")}}
        )
        role = existing["role"]
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "user_id": user_id,
            "email": email,
            "name": oauth_data["name"],
            "role": "coach",
            "picture": oauth_data.get("picture"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user_doc)
        role = "coach"
    
    session_token = oauth_data.get("session_token", f"sess_{uuid.uuid4().hex}")
    session_doc = {
        "session_id": f"sid_{uuid.uuid4().hex[:12]}",
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session_doc)
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    return {
        "user_id": user_id,
        "email": email,
        "name": oauth_data["name"],
        "role": role,
        "picture": oauth_data.get("picture")
    }

@api_router.get("/auth/me")
async def get_me(user: User = Depends(get_current_user)):
    return user.model_dump()

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie("session_token", path="/", samesite="none", secure=True)
    return {"message": "Déconnecté"}

# ============ COMPETITIONS ENDPOINTS ============

async def user_can_access_competition(user: User, competition_id: str) -> bool:
    """Vérifie si l'utilisateur peut accéder à une compétition"""
    # Master et Admin ont accès à toutes les compétitions
    if user.role in ["admin", "master"]:
        return True
    
    competition = await db.competitions.find_one(
        {"competition_id": competition_id},
        {"_id": 0, "coaches_autorises": 1}
    )
    if not competition:
        return False
    
    return user.user_id in competition.get("coaches_autorises", [])

@api_router.get("/competitions")
async def list_competitions(statut: Optional[str] = None, user: User = Depends(get_current_user)):
    """Liste les compétitions accessibles à l'utilisateur"""
    query = {}
    if statut:
        query["statut"] = statut
    
    if user.role in ["admin", "master"]:
        # Admin et Master voient toutes les compétitions
        competitions = await db.competitions.find(query, {"_id": 0}).sort("date", -1).to_list(100)
    else:
        # Coach ne voit que les compétitions où il est autorisé
        query["coaches_autorises"] = user.user_id
        competitions = await db.competitions.find(query, {"_id": 0}).sort("date", -1).to_list(100)
    
    return competitions

@api_router.get("/competitions/{competition_id}")
async def get_competition(competition_id: str, user: User = Depends(get_current_user)):
    """Récupère une compétition spécifique"""
    if not await user_can_access_competition(user, competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé à cette compétition")
    
    competition = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    # Ajouter les stats
    competition["nb_competiteurs"] = await db.competiteurs.count_documents({"competition_id": competition_id})
    competition["nb_combats"] = await db.combats.count_documents({"competition_id": competition_id})
    competition["nb_combats_termines"] = await db.combats.count_documents({"competition_id": competition_id, "termine": True})
    
    return competition

@api_router.post("/competitions")
async def create_competition(data: CompetitionCreate, user: User = Depends(require_admin)):
    """Crée une nouvelle compétition (admin uniquement)"""
    competition = Competition(**data.model_dump(), created_by=user.user_id)
    comp_dict = competition.model_dump()
    comp_dict["created_at"] = comp_dict["created_at"].isoformat()
    
    await db.competitions.insert_one(comp_dict)
    comp_dict.pop("_id", None)
    
    return comp_dict

@api_router.put("/competitions/{competition_id}")
async def update_competition(competition_id: str, data: CompetitionCreate, user: User = Depends(require_admin)):
    """Met à jour une compétition"""
    existing = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    update_data = data.model_dump()
    await db.competitions.update_one(
        {"competition_id": competition_id},
        {"$set": update_data}
    )
    
    return {"message": "Compétition mise à jour"}

@api_router.put("/competitions/{competition_id}/statut")
async def update_competition_statut(competition_id: str, statut: str, user: User = Depends(require_admin)):
    """Change le statut d'une compétition (active, terminee, annulee)"""
    if statut not in ["active", "terminee", "annulee"]:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    result = await db.competitions.update_one(
        {"competition_id": competition_id},
        {"$set": {"statut": statut}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    return {"message": f"Statut mis à jour: {statut}"}

@api_router.delete("/competitions/{competition_id}")
async def delete_competition(competition_id: str, user: User = Depends(require_admin)):
    """Supprime une compétition et toutes ses données"""
    # Supprimer les données liées
    await db.combats.delete_many({"competition_id": competition_id})
    await db.competiteurs.delete_many({"competition_id": competition_id})
    await db.categories.delete_many({"competition_id": competition_id})
    await db.tatamis.delete_many({"competition_id": competition_id})
    await db.aires_combat.delete_many({"competition_id": competition_id})
    await db.medailles.delete_many({"competition_id": competition_id})
    
    result = await db.competitions.delete_one({"competition_id": competition_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    return {"message": "Compétition supprimée"}

@api_router.get("/coaches")
async def list_coaches(user: User = Depends(require_admin)):
    """Liste tous les coachs pour assignation aux compétitions"""
    coaches = await db.users.find({"role": "coach"}, {"_id": 0, "password": 0}).to_list(100)
    return coaches

# ============ COMPETITEURS ENDPOINTS ============

def calculate_age(date_naissance: str) -> int:
    birth = datetime.strptime(date_naissance, "%Y-%m-%d")
    today = datetime.now()
    return today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))

def date_iso_to_fr(date_iso: str) -> str:
    """Convertit une date ISO (YYYY-MM-DD) en format français (JJ/MM/AAAA)"""
    if not date_iso:
        return ""
    try:
        parts = date_iso.split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except:
        pass
    return date_iso

def date_fr_to_iso(date_fr: str) -> str:
    """Convertit une date française (JJ/MM/AAAA) en format ISO (YYYY-MM-DD)"""
    if not date_fr:
        return ""
    try:
        # Gérer les deux formats possibles
        if "/" in str(date_fr):
            parts = str(date_fr).split("/")
            if len(parts) == 3:
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        elif "-" in str(date_fr):
            # Déjà en format ISO
            return str(date_fr)
    except:
        pass
    return str(date_fr)

async def assign_categorie(competiteur: dict, competition_id: str) -> Optional[str]:
    """Assigne une catégorie basée sur le poids officiel (si pesé) ou déclaré"""
    age = calculate_age(competiteur["date_naissance"])
    poids = competiteur.get("poids_officiel") or competiteur.get("poids_declare")
    sexe = competiteur["sexe"]
    
    categorie = await db.categories.find_one({
        "competition_id": competition_id,
        "sexe": sexe,
        "age_min": {"$lte": age},
        "age_max": {"$gte": age},
        "poids_min": {"$lte": poids},
        "poids_max": {"$gte": poids}
    }, {"_id": 0})
    
    return categorie["categorie_id"] if categorie else None

@api_router.get("/competiteurs")
async def list_competiteurs(
    competition_id: Optional[str] = None,
    categorie_id: Optional[str] = None, 
    club: Optional[str] = None,
    pese: Optional[bool] = None,
    user: User = Depends(get_current_user)
):
    """Liste les compétiteurs avec filtres"""
    query = {}
    if competition_id:
        if not await user_can_access_competition(user, competition_id):
            raise HTTPException(status_code=403, detail="Accès non autorisé")
        query["competition_id"] = competition_id
    if categorie_id:
        query["categorie_id"] = categorie_id
    if club:
        query["club"] = club
    if pese is not None:
        query["pese"] = pese
    
    competiteurs = await db.competiteurs.find(query, {"_id": 0}).to_list(1000)
    return competiteurs

@api_router.get("/competiteurs/{competiteur_id}")
async def get_competiteur(competiteur_id: str, user: User = Depends(get_current_user)):
    comp = await db.competiteurs.find_one({"competiteur_id": competiteur_id}, {"_id": 0})
    if not comp:
        raise HTTPException(status_code=404, detail="Compétiteur non trouvé")
    return comp

@api_router.post("/competiteurs")
async def create_competiteur(data: CompetiteurCreate, user: User = Depends(get_current_user)):
    """Crée un compétiteur (coach doit être autorisé sur la compétition)"""
    if not await user_can_access_competition(user, data.competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé à cette compétition")
    
    # Vérifier que la compétition est active
    competition = await db.competitions.find_one({"competition_id": data.competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    if competition.get("statut") != "active" and user.role != "admin":
        raise HTTPException(status_code=400, detail="La compétition n'est plus active")
    
    comp = Competiteur(
        competition_id=data.competition_id,
        nom=data.nom,
        prenom=data.prenom,
        date_naissance=data.date_naissance,
        sexe=data.sexe,
        poids_declare=data.poids_declare,
        club=data.club,
        surclasse=data.surclasse,
        created_by=user.user_id
    )
    comp_dict = comp.model_dump()
    comp_dict["created_at"] = comp_dict["created_at"].isoformat()
    
    # Si surclassé, utiliser la catégorie choisie manuellement
    if data.surclasse and data.categorie_surclasse_id:
        # Vérifier que la catégorie existe et est valide
        categorie = await db.categories.find_one({
            "categorie_id": data.categorie_surclasse_id,
            "competition_id": data.competition_id
        }, {"_id": 0})
        if not categorie:
            raise HTTPException(status_code=400, detail="Catégorie de surclassement invalide")
        
        # Vérifier que le poids est compatible avec la catégorie
        if data.poids_declare < categorie["poids_min"] or data.poids_declare > categorie["poids_max"]:
            raise HTTPException(
                status_code=400, 
                detail=f"Le poids {data.poids_declare}kg n'est pas compatible avec la catégorie {categorie['nom']} ({categorie['poids_min']}-{categorie['poids_max']}kg)"
            )
        
        comp_dict["categorie_id"] = data.categorie_surclasse_id
    else:
        # Attribution automatique basée sur l'âge et le poids
        categorie_id = await assign_categorie(comp_dict, data.competition_id)
        comp_dict["categorie_id"] = categorie_id
    
    await db.competiteurs.insert_one(comp_dict)
    comp_dict.pop("_id", None)
    return comp_dict

@api_router.put("/competiteurs/{competiteur_id}")
async def update_competiteur(competiteur_id: str, data: CompetiteurCreate, user: User = Depends(require_admin)):
    existing = await db.competiteurs.find_one({"competiteur_id": competiteur_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Compétiteur non trouvé")
    
    update_data = data.model_dump()
    categorie_id = await assign_categorie({**update_data, "date_naissance": update_data["date_naissance"]}, data.competition_id)
    update_data["categorie_id"] = categorie_id
    
    await db.competiteurs.update_one(
        {"competiteur_id": competiteur_id},
        {"$set": update_data}
    )
    
    updated = await db.competiteurs.find_one({"competiteur_id": competiteur_id}, {"_id": 0})
    return updated

@api_router.delete("/competiteurs/{competiteur_id}")
async def delete_competiteur(competiteur_id: str, user: User = Depends(require_admin)):
    result = await db.competiteurs.delete_one({"competiteur_id": competiteur_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Compétiteur non trouvé")
    return {"message": "Compétiteur supprimé"}

# ============ PESEE ENDPOINTS ============

@api_router.get("/pesee/{competition_id}")
async def list_pesee(competition_id: str, user: User = Depends(get_current_user)):
    """Liste tous les compétiteurs pour la pesée d'une compétition"""
    if not await user_can_access_competition(user, competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    competiteurs = await db.competiteurs.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).sort([("club", 1), ("nom", 1)]).to_list(1000)
    
    # Enrichir avec les noms des catégories
    for comp in competiteurs:
        if comp.get("categorie_id"):
            cat = await db.categories.find_one({"categorie_id": comp["categorie_id"]}, {"_id": 0, "nom": 1})
            comp["categorie_nom"] = cat["nom"] if cat else "Non assignée"
        else:
            comp["categorie_nom"] = "Non assignée"
    
    return competiteurs

@api_router.put("/pesee/{competiteur_id}")
async def enregistrer_pesee(competiteur_id: str, data: PeseeUpdate, user: User = Depends(require_admin)):
    """Enregistre le poids officiel d'un compétiteur (admin uniquement)"""
    comp = await db.competiteurs.find_one({"competiteur_id": competiteur_id}, {"_id": 0})
    if not comp:
        raise HTTPException(status_code=404, detail="Compétiteur non trouvé")
    
    # Mettre à jour le poids officiel
    update_data = {
        "poids_officiel": data.poids_officiel,
        "pese": True
    }
    
    # Recalculer la catégorie basée sur le poids officiel
    comp_updated = {**comp, "poids_officiel": data.poids_officiel}
    nouvelle_categorie = await assign_categorie(comp_updated, comp["competition_id"])
    
    ancienne_categorie = comp.get("categorie_id")
    update_data["categorie_id"] = nouvelle_categorie
    
    await db.competiteurs.update_one(
        {"competiteur_id": competiteur_id},
        {"$set": update_data}
    )
    
    # Si la catégorie a changé, notifier
    categorie_changee = ancienne_categorie != nouvelle_categorie
    
    return {
        "message": "Pesée enregistrée",
        "poids_officiel": data.poids_officiel,
        "categorie_id": nouvelle_categorie,
        "categorie_changee": categorie_changee
    }

@api_router.put("/pesee/{competiteur_id}/poids-declare")
async def enregistrer_poids_declare(competiteur_id: str, poids: float, user: User = Depends(get_current_user)):
    """Enregistre/modifie le poids déclaré (coach peut modifier)"""
    comp = await db.competiteurs.find_one({"competiteur_id": competiteur_id}, {"_id": 0})
    if not comp:
        raise HTTPException(status_code=404, detail="Compétiteur non trouvé")
    
    if not await user_can_access_competition(user, comp["competition_id"]):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    await db.competiteurs.update_one(
        {"competiteur_id": competiteur_id},
        {"$set": {"poids_declare": poids}}
    )
    
    return {"message": "Poids déclaré mis à jour", "poids_declare": poids}

@api_router.delete("/pesee/{competiteur_id}")
async def annuler_pesee(competiteur_id: str, user: User = Depends(require_admin)):
    """Annule la pesée d'un compétiteur"""
    comp = await db.competiteurs.find_one({"competiteur_id": competiteur_id}, {"_id": 0})
    if not comp:
        raise HTTPException(status_code=404, detail="Compétiteur non trouvé")
    
    # Recalculer la catégorie basée sur le poids déclaré
    comp_updated = {**comp, "poids_officiel": None}
    nouvelle_categorie = await assign_categorie(comp_updated, comp["competition_id"])
    
    await db.competiteurs.update_one(
        {"competiteur_id": competiteur_id},
        {"$set": {"poids_officiel": None, "pese": False, "categorie_id": nouvelle_categorie}}
    )
    
    return {"message": "Pesée annulée"}

# ============ CATEGORIES ENDPOINTS ============

@api_router.get("/categories")
async def list_categories(competition_id: Optional[str] = None, user: User = Depends(get_current_user)):
    query = {}
    if competition_id:
        query["competition_id"] = competition_id
    
    categories = await db.categories.find(query, {"_id": 0}).to_list(500)
    return categories

@api_router.post("/categories")
async def create_categorie(data: CategorieCreate, user: User = Depends(require_admin)):
    cat = Categorie(**data.model_dump())
    cat_dict = cat.model_dump()
    await db.categories.insert_one(cat_dict)
    cat_dict.pop("_id", None)
    return cat_dict

@api_router.delete("/categories/{categorie_id}")
async def delete_categorie(categorie_id: str, user: User = Depends(require_admin)):
    result = await db.categories.delete_one({"categorie_id": categorie_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Catégorie non trouvée")
    return {"message": "Catégorie supprimée"}

# ============ SEED CATEGORIES OFFICIELLES ============

# Définition des catégories officielles FFTA/FFDA 2025/2026
CATEGORIES_OFFICIELLES = {
    "Pupilles 1": {
        "age_min": 6, "age_max": 6,
        "M": [("-21kg", 0, 21), ("-24kg", 21, 24), ("-27kg", 24, 27), ("+27kg", 27, 200)],
        "F": [("-17kg", 0, 17), ("-20kg", 17, 20), ("-23kg", 20, 23), ("+23kg", 23, 200)]
    },
    "Pupilles 2": {
        "age_min": 7, "age_max": 7,
        "M": [("-21kg", 0, 21), ("-24kg", 21, 24), ("-27kg", 24, 27), ("-30kg", 27, 30), ("+30kg", 30, 200)],
        "F": [("-17kg", 0, 17), ("-20kg", 17, 20), ("-23kg", 20, 23), ("-26kg", 23, 26), ("+26kg", 26, 200)]
    },
    "Benjamins": {
        "age_min": 8, "age_max": 9,
        "M": [("-21kg", 0, 21), ("-24kg", 21, 24), ("-27kg", 24, 27), ("-30kg", 27, 30), ("-33kg", 30, 33), 
              ("-37kg", 33, 37), ("-41kg", 37, 41), ("-45kg", 41, 45), ("-49kg", 45, 49), ("+49kg", 49, 200)],
        "F": [("-17kg", 0, 17), ("-20kg", 17, 20), ("-23kg", 20, 23), ("-26kg", 23, 26), ("-29kg", 26, 29), 
              ("-33kg", 29, 33), ("-37kg", 33, 37), ("-41kg", 37, 41), ("-44kg", 41, 44), ("+44kg", 44, 200)]
    },
    "Minimes": {
        "age_min": 10, "age_max": 11,
        "M": [("-27kg", 0, 27), ("-30kg", 27, 30), ("-33kg", 30, 33), ("-37kg", 33, 37), ("-41kg", 37, 41), 
              ("-45kg", 41, 45), ("-49kg", 45, 49), ("-53kg", 49, 53), ("-57kg", 53, 57), ("+57kg", 57, 200)],
        "F": [("-23kg", 0, 23), ("-26kg", 23, 26), ("-29kg", 26, 29), ("-33kg", 29, 33), ("-37kg", 33, 37), 
              ("-41kg", 37, 41), ("-44kg", 41, 44), ("-47kg", 44, 47), ("-51kg", 47, 51), ("+51kg", 51, 200)]
    },
    "Cadets": {
        "age_min": 12, "age_max": 13,
        "M": [("-33kg", 0, 33), ("-37kg", 33, 37), ("-41kg", 37, 41), ("-45kg", 41, 45), ("-49kg", 45, 49), 
              ("-53kg", 49, 53), ("-57kg", 53, 57), ("-61kg", 57, 61), ("-65kg", 61, 65), ("+65kg", 65, 200)],
        "F": [("-29kg", 0, 29), ("-33kg", 29, 33), ("-37kg", 33, 37), ("-41kg", 37, 41), ("-44kg", 41, 44), 
              ("-47kg", 44, 47), ("-51kg", 47, 51), ("-55kg", 51, 55), ("-59kg", 55, 59), ("+59kg", 59, 200)]
    },
    "Juniors": {
        "age_min": 14, "age_max": 17,
        "M": [("-45kg", 0, 45), ("-48kg", 45, 48), ("-51kg", 48, 51), ("-55kg", 51, 55), ("-59kg", 55, 59), 
              ("-63kg", 59, 63), ("-68kg", 63, 68), ("-73kg", 68, 73), ("-78kg", 73, 78), ("+78kg", 78, 200)],
        "F": [("-42kg", 0, 42), ("-44kg", 42, 44), ("-46kg", 44, 46), ("-49kg", 46, 49), ("-52kg", 49, 52), 
              ("-55kg", 52, 55), ("-59kg", 55, 59), ("-63kg", 59, 63), ("-68kg", 63, 68), ("+68kg", 68, 200)]
    },
    "Seniors": {
        "age_min": 18, "age_max": 29,
        "M": [("-54kg", 0, 54), ("-58kg", 54, 58), ("-63kg", 58, 63), ("-68kg", 63, 68), ("-74kg", 68, 74), 
              ("-80kg", 74, 80), ("-87kg", 80, 87), ("+87kg", 87, 200)],
        "F": [("-46kg", 0, 46), ("-49kg", 46, 49), ("-53kg", 49, 53), ("-57kg", 53, 57), ("-62kg", 57, 62), 
              ("-67kg", 62, 67), ("-73kg", 67, 73), ("+73kg", 73, 200)]
    },
    "Masters": {
        "age_min": 30, "age_max": 99,
        "M": [("-58kg", 0, 58), ("-63kg", 58, 63), ("-68kg", 63, 68), ("-74kg", 68, 74), ("-80kg", 74, 80), ("+80kg", 80, 200)],
        "F": [("-49kg", 0, 49), ("-53kg", 49, 53), ("-57kg", 53, 57), ("-62kg", 57, 62), ("-67kg", 62, 67), ("+67kg", 67, 200)]
    }
}

@api_router.post("/categories/seed/{competition_id}")
async def seed_categories(competition_id: str, user: User = Depends(require_admin)):
    """Peuple une compétition avec toutes les catégories officielles FFTA/FFDA"""
    # Vérifier que la compétition existe
    competition = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    # Supprimer les catégories existantes de cette compétition
    await db.categories.delete_many({"competition_id": competition_id})
    
    categories_created = []
    
    for categorie_age, config in CATEGORIES_OFFICIELLES.items():
        age_min = config["age_min"]
        age_max = config["age_max"]
        
        # Catégories masculines
        for poids_nom, poids_min, poids_max in config["M"]:
            cat = Categorie(
                competition_id=competition_id,
                nom=f"{categorie_age} Masculin {poids_nom}",
                age_min=age_min,
                age_max=age_max,
                sexe="M",
                poids_min=poids_min,
                poids_max=poids_max
            )
            cat_dict = cat.model_dump()
            await db.categories.insert_one(cat_dict)
            cat_dict.pop("_id", None)
            categories_created.append(cat_dict)
        
        # Catégories féminines
        for poids_nom, poids_min, poids_max in config["F"]:
            cat = Categorie(
                competition_id=competition_id,
                nom=f"{categorie_age} Féminin {poids_nom}",
                age_min=age_min,
                age_max=age_max,
                sexe="F",
                poids_min=poids_min,
                poids_max=poids_max
            )
            cat_dict = cat.model_dump()
            await db.categories.insert_one(cat_dict)
            cat_dict.pop("_id", None)
            categories_created.append(cat_dict)
    
    return {
        "message": f"{len(categories_created)} catégories créées pour la compétition",
        "total": len(categories_created)
    }

@api_router.get("/categories/age-groups")
async def get_age_groups():
    """Retourne la liste des groupes d'âge pour le surclassement"""
    groups = []
    for nom, config in CATEGORIES_OFFICIELLES.items():
        groups.append({
            "nom": nom,
            "age_min": config["age_min"],
            "age_max": config["age_max"]
        })
    return groups

@api_router.get("/categories/for-surclassement/{competition_id}")
async def get_categories_for_surclassement(
    competition_id: str,
    sexe: str,
    age: int,
    user: User = Depends(get_current_user)
):
    """Retourne les catégories disponibles pour le surclassement (âge égal ou supérieur)"""
    if not await user_can_access_competition(user, competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    # Récupérer toutes les catégories du sexe demandé avec âge >= âge du compétiteur
    categories = await db.categories.find({
        "competition_id": competition_id,
        "sexe": sexe,
        "age_min": {"$gte": age}
    }, {"_id": 0}).sort([("age_min", 1), ("poids_min", 1)]).to_list(500)
    
    return categories

# ============ AIRES DE COMBAT ENDPOINTS ============

@api_router.get("/aires-combat")
async def list_aires_combat(competition_id: Optional[str] = None, user: User = Depends(get_current_user)):
    """Liste les aires de combat d'une compétition"""
    query = {}
    if competition_id:
        query["competition_id"] = competition_id
    
    aires = await db.aires_combat.find(query, {"_id": 0}).sort("numero", 1).to_list(100)
    # Assurer que chaque aire a un statut (pour rétrocompatibilité)
    for aire in aires:
        if "statut" not in aire:
            aire["statut"] = "active"
    return aires

@api_router.get("/aires-combat/{aire_id}")
async def get_aire_combat(aire_id: str, user: User = Depends(get_current_user)):
    """Récupère une aire de combat avec ses combats en cours"""
    aire = await db.aires_combat.find_one({"aire_id": aire_id}, {"_id": 0})
    if not aire:
        raise HTTPException(status_code=404, detail="Aire de combat non trouvée")
    
    # Assurer que l'aire a un statut (pour rétrocompatibilité)
    if "statut" not in aire:
        aire["statut"] = "active"
    
    # Récupérer le combat en cours sur cette aire
    combat_en_cours = await db.combats.find_one(
        {"aire_id": aire_id, "statut": "en_cours"},
        {"_id": 0}
    )
    
    # Récupérer les combats à venir sur cette aire
    combats_a_venir = await db.combats.find(
        {"aire_id": aire_id, "statut": "a_venir", "est_finale": False},
        {"_id": 0}
    ).sort("ordre", 1).to_list(20)
    
    aire["combat_en_cours"] = combat_en_cours
    aire["combats_a_venir"] = combats_a_venir
    
    return aire

@api_router.post("/aires-combat")
async def create_aire_combat(data: AireCombatCreate, user: User = Depends(require_admin)):
    """Crée une nouvelle aire de combat"""
    aire = AireCombat(**data.model_dump())
    aire_dict = aire.model_dump()
    await db.aires_combat.insert_one(aire_dict)
    aire_dict.pop("_id", None)
    return aire_dict

@api_router.delete("/aires-combat/{aire_id}")
async def delete_aire_combat(aire_id: str, user: User = Depends(require_admin)):
    """Supprime une aire de combat"""
    result = await db.aires_combat.delete_one({"aire_id": aire_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Aire de combat non trouvée")
    return {"message": "Aire de combat supprimée"}

@api_router.post("/aires-combat/repartir/{competition_id}")
async def repartir_combats_sur_aires(competition_id: str, user: User = Depends(require_admin)):
    """
    Répartit automatiquement les combats sur les aires de combat disponibles.
    Les finales sont mises à la fin de la compétition.
    """
    # Récupérer les aires de combat
    aires = await db.aires_combat.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).sort("numero", 1).to_list(10)
    
    if not aires:
        raise HTTPException(status_code=400, detail="Aucune aire de combat configurée")
    
    # Récupérer tous les combats non terminés
    combats = await db.combats.find(
        {"competition_id": competition_id, "termine": False},
        {"_id": 0}
    ).to_list(1000)
    
    if not combats:
        return {"message": "Aucun combat à répartir", "total": 0}
    
    # Séparer finales et autres combats
    finales = [c for c in combats if c["tour"] in ["finale", "bronze"]]
    autres = [c for c in combats if c["tour"] not in ["finale", "bronze"]]
    
    # Trier par catégorie puis par tour
    tour_ordre = {"quart": 1, "demi": 2}
    autres.sort(key=lambda x: (x["categorie_id"], tour_ordre.get(x["tour"], 0), x["position"]))
    finales.sort(key=lambda x: (x["categorie_id"], 1 if x["tour"] == "bronze" else 2, x["position"]))
    
    # Répartir les combats (sauf finales) sur les aires
    nb_aires = len(aires)
    for i, combat in enumerate(autres):
        aire = aires[i % nb_aires]
        await db.combats.update_one(
            {"combat_id": combat["combat_id"]},
            {"$set": {
                "aire_id": aire["aire_id"],
                "ordre": i + 1,
                "est_finale": False
            }}
        )
    
    # Les finales seront sur toutes les aires (rotation) à la fin
    base_ordre = len(autres)
    for i, combat in enumerate(finales):
        aire = aires[i % nb_aires]
        await db.combats.update_one(
            {"combat_id": combat["combat_id"]},
            {"$set": {
                "aire_id": aire["aire_id"],
                "ordre": base_ordre + i + 1,
                "est_finale": True
            }}
        )
    
    return {
        "message": f"{len(combats)} combats répartis sur {nb_aires} aire(s)",
        "combats_reguliers": len(autres),
        "finales": len(finales)
    }

@api_router.put("/aires-combat/{aire_id}")
async def update_aire_combat(aire_id: str, data: AireCombatUpdate, user: User = Depends(require_admin)):
    """Met à jour une aire de combat (nom, statut: active/pause/hs)"""
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    
    result = await db.aires_combat.update_one(
        {"aire_id": aire_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Aire de combat non trouvée")
    
    aire = await db.aires_combat.find_one({"aire_id": aire_id}, {"_id": 0})
    return aire

# ============ GESTION ORDRE DES COMBATS (DRAG & DROP) ============

class ReorderCombatsRequest(BaseModel):
    combat_ids: list[str]  # Liste des combat_ids dans le nouvel ordre

@api_router.put("/combats/reorder/{aire_id}")
async def reorder_combats(aire_id: str, data: ReorderCombatsRequest, user: User = Depends(require_admin)):
    """
    Réordonne les combats d'une aire de combat.
    Reçoit la liste des combat_ids dans le nouvel ordre.
    """
    # Vérifier que l'aire existe
    aire = await db.aires_combat.find_one({"aire_id": aire_id}, {"_id": 0})
    if not aire:
        raise HTTPException(status_code=404, detail="Aire de combat non trouvée")
    
    # Mettre à jour l'ordre de chaque combat
    for index, combat_id in enumerate(data.combat_ids):
        await db.combats.update_one(
            {"combat_id": combat_id, "aire_id": aire_id},
            {"$set": {"ordre": index + 1}}
        )
    
    return {"message": f"{len(data.combat_ids)} combat(s) réordonnés"}

@api_router.get("/combats/ordre/{aire_id}")
async def get_combats_ordre(aire_id: str, user: User = Depends(get_current_user)):
    """Récupère les combats d'une aire dans l'ordre"""
    combats = await db.combats.find(
        {"aire_id": aire_id, "termine": False},
        {"_id": 0}
    ).sort([("est_finale", 1), ("ordre", 1)]).to_list(200)
    
    # ============ OPTIMISATION: Batch fetch pour éviter N+1 queries ============
    # Collecter tous les IDs uniques
    competiteur_ids = set()
    categorie_ids = set()
    for combat in combats:
        if combat.get("rouge_id"):
            competiteur_ids.add(combat["rouge_id"])
        if combat.get("bleu_id"):
            competiteur_ids.add(combat["bleu_id"])
        if combat.get("categorie_id"):
            categorie_ids.add(combat["categorie_id"])
    
    # Fetch en batch
    competiteurs_list = await db.competiteurs.find(
        {"competiteur_id": {"$in": list(competiteur_ids)}},
        {"_id": 0, "competiteur_id": 1, "nom": 1, "prenom": 1, "club": 1}
    ).to_list(500)
    competiteurs_map = {c["competiteur_id"]: c for c in competiteurs_list}
    
    categories_list = await db.categories.find(
        {"categorie_id": {"$in": list(categorie_ids)}},
        {"_id": 0, "categorie_id": 1, "nom": 1}
    ).to_list(200)
    categories_map = {c["categorie_id"]: c for c in categories_list}
    
    # Enrichir avec les infos (lookup rapide)
    for combat in combats:
        if combat.get("rouge_id"):
            combat["rouge"] = competiteurs_map.get(combat["rouge_id"])
        if combat.get("bleu_id"):
            combat["bleu"] = competiteurs_map.get(combat["bleu_id"])
        if combat.get("categorie_id"):
            combat["categorie"] = categories_map.get(combat["categorie_id"])
    
    return combats

# ============ GESTION FORFAITS / ABSENCES ============

class ForfaitRequest(BaseModel):
    competiteur_id: str
    raison: str = "forfait"  # forfait, absence, blessure, disqualification

@api_router.post("/combats/{combat_id}/forfait")
async def declarer_forfait(combat_id: str, data: ForfaitRequest, user: User = Depends(get_current_user)):
    """
    Déclare un forfait pour un combattant.
    Le combat est automatiquement gagné par l'adversaire.
    """
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    if combat["termine"]:
        raise HTTPException(status_code=400, detail="Ce combat est déjà terminé")
    
    # Déterminer qui fait forfait et qui gagne
    if data.competiteur_id == combat.get("rouge_id"):
        vainqueur_id = combat.get("bleu_id")
        forfait_couleur = "rouge"
    elif data.competiteur_id == combat.get("bleu_id"):
        vainqueur_id = combat.get("rouge_id")
        forfait_couleur = "bleu"
    else:
        raise HTTPException(status_code=400, detail="Ce compétiteur n'est pas dans ce combat")
    
    # Mettre à jour le combat
    update_data = {
        "type_victoire": data.raison,
        "termine": True,
        "statut": "termine"
    }
    
    if vainqueur_id:
        update_data["vainqueur_id"] = vainqueur_id
    else:
        # Si l'adversaire n'est pas encore défini, le combat est annulé
        update_data["statut"] = "non_dispute"
    
    await db.combats.update_one(
        {"combat_id": combat_id},
        {"$set": update_data}
    )
    
    # Marquer le compétiteur comme éliminé
    await db.competiteurs.update_one(
        {"competiteur_id": data.competiteur_id},
        {"$set": {"elimine": True, "raison_elimination": data.raison}}
    )
    
    # Propager le vainqueur si défini
    if vainqueur_id:
        await propager_vainqueur(combat, vainqueur_id)
    
    # Log de l'action
    await db.historique_resultats.insert_one({
        "historique_id": f"hist_{uuid.uuid4().hex[:12]}",
        "combat_id": combat_id,
        "action": "forfait",
        "competiteur_id": data.competiteur_id,
        "raison": data.raison,
        "modifie_par": user.user_id,
        "date": datetime.now(timezone.utc).isoformat()
    })
    
    updated = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    return {
        "message": f"Forfait enregistré ({data.raison})",
        "combat": updated,
        "vainqueur_id": vainqueur_id
    }

# ============ TATAMIS ENDPOINTS (RETRO-COMPATIBILITE) ============

@api_router.get("/tatamis")
async def list_tatamis(competition_id: Optional[str] = None, user: User = Depends(get_current_user)):
    query = {}
    if competition_id:
        query["competition_id"] = competition_id
    
    tatamis = await db.tatamis.find(query, {"_id": 0}).to_list(100)
    return tatamis

@api_router.post("/tatamis")
async def create_tatami(data: TatamiCreate, user: User = Depends(require_admin)):
    tat = Tatami(**data.model_dump())
    tat_dict = tat.model_dump()
    await db.tatamis.insert_one(tat_dict)
    tat_dict.pop("_id", None)
    return tat_dict

@api_router.delete("/tatamis/{tatami_id}")
async def delete_tatami(tatami_id: str, user: User = Depends(require_admin)):
    result = await db.tatamis.delete_one({"tatami_id": tatami_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tatami non trouvé")
    return {"message": "Tatami supprimé"}

# ============ COMBATS ENDPOINTS ============

@api_router.get("/combats/suivre")
async def combats_a_suivre(
    competition_id: Optional[str] = None,
    categorie_id: Optional[str] = None,
    tatami_id: Optional[str] = None,
    tour: Optional[str] = None,
    statut: Optional[str] = None,
    user: User = Depends(get_current_user)
):
    """Récupère les combats à suivre avec filtres"""
    query = {"statut": {"$ne": "termine"}} if not statut else {}
    
    if competition_id:
        query["competition_id"] = competition_id
    if categorie_id:
        query["categorie_id"] = categorie_id
    if tatami_id:
        query["tatami_id"] = tatami_id
    if tour:
        query["tour"] = tour
    if statut:
        query["statut"] = statut
    
    combats = await db.combats.find(query, {"_id": 0}).sort("ordre", 1).to_list(500)
    
    # ============ OPTIMISATION: Batch fetch pour éviter N+1 queries ============
    competiteur_ids = set()
    categorie_ids = set()
    tatami_ids = set()
    for combat in combats:
        if combat.get("rouge_id"):
            competiteur_ids.add(combat["rouge_id"])
        if combat.get("bleu_id"):
            competiteur_ids.add(combat["bleu_id"])
        if combat.get("categorie_id"):
            categorie_ids.add(combat["categorie_id"])
        if combat.get("tatami_id"):
            tatami_ids.add(combat["tatami_id"])
    
    competiteurs_list = await db.competiteurs.find(
        {"competiteur_id": {"$in": list(competiteur_ids)}},
        {"_id": 0, "competiteur_id": 1, "nom": 1, "prenom": 1, "club": 1}
    ).to_list(1000)
    competiteurs_map = {c["competiteur_id"]: c for c in competiteurs_list}
    
    categories_list = await db.categories.find(
        {"categorie_id": {"$in": list(categorie_ids)}},
        {"_id": 0, "categorie_id": 1, "nom": 1}
    ).to_list(200)
    categories_map = {c["categorie_id"]: c for c in categories_list}
    
    tatamis_list = await db.tatamis.find(
        {"tatami_id": {"$in": list(tatami_ids)}},
        {"_id": 0, "tatami_id": 1, "nom": 1, "numero": 1}
    ).to_list(50)
    tatamis_map = {t["tatami_id"]: t for t in tatamis_list}
    
    # Enrichir avec les noms (lookup rapide)
    for combat in combats:
        if combat.get("rouge_id"):
            rouge = competiteurs_map.get(combat["rouge_id"])
            combat["rouge_nom"] = f"{rouge['prenom']} {rouge['nom']}" if rouge else "Inconnu"
            combat["rouge_club"] = rouge.get("club", "") if rouge else ""
        else:
            combat["rouge_nom"] = "À déterminer"
            combat["rouge_club"] = ""
            
        if combat.get("bleu_id"):
            bleu = competiteurs_map.get(combat["bleu_id"])
            combat["bleu_nom"] = f"{bleu['prenom']} {bleu['nom']}" if bleu else "Inconnu"
            combat["bleu_club"] = bleu.get("club", "") if bleu else ""
        else:
            combat["bleu_nom"] = "À déterminer"
            combat["bleu_club"] = ""
        
        cat = categories_map.get(combat.get("categorie_id"))
        combat["categorie_nom"] = cat["nom"] if cat else "Inconnue"
        
        if combat.get("tatami_id"):
            tatami = tatamis_map.get(combat["tatami_id"])
            combat["tatami_nom"] = tatami["nom"] if tatami else "Non assigné"
        else:
            combat["tatami_nom"] = "Non assigné"
    
    return combats

@api_router.get("/combats/arbre/{categorie_id}")
async def get_arbre_combats(categorie_id: str, user: User = Depends(get_current_user)):
    """
    Récupère l'arbre complet des combats pour une catégorie (pour affichage et export PDF).
    Inclut les tours: seizieme, huitieme, quart, demi, finale.
    PAS de combat bronze (règle World Taekwondo).
    """
    combats = await db.combats.find({"categorie_id": categorie_id}, {"_id": 0}).to_list(100)
    
    # ============ OPTIMISATION: Batch fetch pour éviter N+1 queries ============
    competiteur_ids = set()
    for combat in combats:
        if combat.get("rouge_id"):
            competiteur_ids.add(combat["rouge_id"])
        if combat.get("bleu_id"):
            competiteur_ids.add(combat["bleu_id"])
        if combat.get("vainqueur_id"):
            competiteur_ids.add(combat["vainqueur_id"])
    
    competiteurs_list = await db.competiteurs.find(
        {"competiteur_id": {"$in": list(competiteur_ids)}},
        {"_id": 0, "competiteur_id": 1, "nom": 1, "prenom": 1, "club": 1}
    ).to_list(500)
    competiteurs_map = {c["competiteur_id"]: c for c in competiteurs_list}
    
    # Structure dynamique pour tous les tours possibles
    arbre = {"seizieme": [], "huitieme": [], "quart": [], "demi": [], "finale": []}
    
    for combat in combats:
        # Ajouter les noms des compétiteurs (lookup rapide)
        if combat.get("rouge_id"):
            rouge = competiteurs_map.get(combat["rouge_id"])
            combat["rouge"] = {"nom": f"{rouge['prenom']} {rouge['nom']}", "club": rouge.get("club", "")} if rouge else {"nom": "Inconnu", "club": ""}
        else:
            combat["rouge"] = {"nom": "À déterminer", "club": ""}
            
        if combat.get("bleu_id"):
            bleu = competiteurs_map.get(combat["bleu_id"])
            combat["bleu"] = {"nom": f"{bleu['prenom']} {bleu['nom']}", "club": bleu.get("club", "")} if bleu else {"nom": "Inconnu", "club": ""}
        else:
            combat["bleu"] = {"nom": "À déterminer", "club": ""}
        
        # Ajouter le vainqueur si terminé
        if combat.get("vainqueur_id"):
            vainqueur = competiteurs_map.get(combat["vainqueur_id"])
            combat["vainqueur_nom"] = f"{vainqueur['prenom']} {vainqueur['nom']}" if vainqueur else "Inconnu"
        
        # Ajouter au bon tour (ignorer les combats bronze qui n'existent plus)
        tour = combat.get("tour", "")
        if tour in arbre:
            arbre[tour].append(combat)
        elif tour and tour not in ["bronze"]:
            # Tour personnalisé ou inconnu
            if tour not in arbre:
                arbre[tour] = []
            arbre[tour].append(combat)
    
    # Trier par position
    for tour in arbre:
        arbre[tour] = sorted(arbre[tour], key=lambda x: x.get("position", 0))
    
    # Supprimer les tours vides
    arbre = {k: v for k, v in arbre.items() if v}
    
    # Ajouter les infos de la catégorie
    categorie = await db.categories.find_one({"categorie_id": categorie_id}, {"_id": 0})
    
    # Calculer le bracket_size
    nb_competiteurs = categorie.get("nb_combattants", 0) if categorie else 0
    
    return {
        "categorie": categorie,
        "arbre": arbre,
        "total_combats": len(combats),
        "combats_termines": len([c for c in combats if c.get("termine")]),
        "bracket_size": categorie.get("bracket_size", 0) if categorie else 0,
        "num_byes": categorie.get("num_byes", 0) if categorie else 0,
        "byes_locked": categorie.get("byes_locked", False) if categorie else False
    }

@api_router.get("/combats")
async def list_combats(
    competition_id: Optional[str] = None,
    categorie_id: Optional[str] = None,
    tatami_id: Optional[str] = None,
    tour: Optional[str] = None,
    user: User = Depends(get_current_user)
):
    query = {}
    if competition_id:
        query["competition_id"] = competition_id
    if categorie_id:
        query["categorie_id"] = categorie_id
    if tatami_id:
        query["tatami_id"] = tatami_id
    if tour:
        query["tour"] = tour
    
    combats = await db.combats.find(query, {"_id": 0}).to_list(1000)
    return combats

@api_router.get("/combats/{combat_id}")
async def get_combat(combat_id: str, user: User = Depends(get_current_user)):
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    return combat

@api_router.post("/combats/generer/{categorie_id}")
async def generer_tableau(categorie_id: str, tatami_id: Optional[str] = None, user: User = Depends(require_admin)):
    """
    Génère l'arbre des combats pour une catégorie selon les règles World Taekwondo.
    
    RÈGLE FONDAMENTALE - PUISSANCE DE 2:
    Le bracket est TOUJOURS basé sur la puissance de 2 immédiatement supérieure.
    - 2 combattants  → bracket de 2  (finale)
    - 3-4 combattants  → bracket de 4  (demis + finale)
    - 5-8 combattants  → bracket de 8  (quarts + demis + finale)
    - 9-16 combattants → bracket de 16 (8èmes + quarts + demis + finale)
    - 17-32 combattants → bracket de 32 (16èmes + 8èmes + quarts + demis + finale)
    
    JAMAIS de bracket intermédiaire, JAMAIS de tours manquants.
    
    PAS de combat bronze (règle World Taekwondo) - 2 bronze ex-aequo.
    """
    # Récupérer la catégorie
    categorie = await db.categories.find_one({"categorie_id": categorie_id}, {"_id": 0})
    if not categorie:
        raise HTTPException(status_code=404, detail="Catégorie non trouvée")
    
    competition_id = categorie.get("competition_id")
    
    # Supprimer les anciens combats de cette catégorie
    await db.combats.delete_many({"categorie_id": categorie_id})
    await db.medailles.delete_many({"categorie_id": categorie_id})
    
    # Récupérer les compétiteurs (non disqualifiés, triés par inscription)
    competiteurs = await db.competiteurs.find(
        {"categorie_id": categorie_id, "disqualifie": False},
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    
    n = len(competiteurs)
    
    if n == 0:
        raise HTTPException(status_code=400, detail="Aucun compétiteur dans cette catégorie")
    
    if n == 1:
        raise HTTPException(
            status_code=400, 
            detail="⚠ Catégorie incomplète — 1 seul participant inscrit. Il faut au moins 2 compétiteurs."
        )
    
    # ============ CALCUL DU BRACKET ============
    def next_power_of_2(x):
        """Retourne la puissance de 2 >= x"""
        if x <= 1:
            return 2
        return 1 << (x - 1).bit_length()
    
    bracket_size = next_power_of_2(n)
    num_byes = bracket_size - n
    
    # Nombre de tours = log2(bracket_size)
    import math
    num_rounds = int(math.log2(bracket_size))
    
    # Noms des tours selon le bracket_size
    def get_tour_names(bracket_size):
        """Retourne la liste des tours dans l'ordre du premier au dernier"""
        tours = []
        if bracket_size >= 32:
            tours.append("seizieme")
        if bracket_size >= 16:
            tours.append("huitieme")
        if bracket_size >= 8:
            tours.append("quart")
        if bracket_size >= 4:
            tours.append("demi")
        tours.append("finale")
        return tours
    
    tour_names = get_tour_names(bracket_size)
    
    # ============ MÉLANGER ET DISTRIBUER LES BYES ============
    random.shuffle(competiteurs)
    
    # Créer les slots du premier tour (bracket_size positions)
    # Les BYEs sont distribués stratégiquement pour éviter que 2 BYEs se rencontrent
    # Positions idéales pour les BYEs: positions impaires (1, 3, 5, 7...) espacées uniformément
    slots = [None] * bracket_size
    
    # Calculer les positions de BYE optimales
    bye_positions = set()
    if num_byes > 0:
        # Espacer les BYEs uniformément dans le bracket
        # Utiliser les positions impaires en priorité pour une meilleure répartition
        step = bracket_size / num_byes
        for i in range(num_byes):
            # Position de base + décalage pour éviter les positions paires
            pos = int(i * step)
            # S'assurer que ce n'est pas une position déjà prise et que le BYE est en position "basse" du combat
            # (position impaire = combattant 2 du combat)
            if pos % 2 == 0:
                pos = pos + 1 if pos + 1 < bracket_size else pos - 1
            while pos in bye_positions or pos >= bracket_size:
                pos = (pos + 2) % bracket_size
            bye_positions.add(pos)
    
    # Remplir les slots avec les compétiteurs
    comp_idx = 0
    for i in range(bracket_size):
        if i in bye_positions:
            slots[i] = None  # BYE
        else:
            if comp_idx < len(competiteurs):
                slots[i] = competiteurs[comp_idx]
                comp_idx += 1
            else:
                slots[i] = None  # BYE supplémentaire
    
    # ============ GÉNÉRER TOUS LES COMBATS DE TOUS LES TOURS ============
    combats_created = []
    combats_by_tour = {tour: [] for tour in tour_names}
    
    # Structure pour suivre les combats par position dans chaque tour
    # combat_map[tour][position] = combat_id
    combat_map = {tour: {} for tour in tour_names}
    
    async def insert_combat(competition_id, categorie_id, tatami_id, tour, position, rouge_id, bleu_id, pret=False):
        """Crée et insère un combat, retourne le dict sans _id"""
        combat = Combat(
            competition_id=competition_id,
            categorie_id=categorie_id,
            tatami_id=tatami_id,
            tour=tour,
            position=position,
            rouge_id=rouge_id,
            bleu_id=bleu_id,
            has_bye=False,
            pret=pret
        )
        combat_dict = combat.model_dump()
        combat_dict["created_at"] = combat_dict["created_at"].isoformat()
        await db.combats.insert_one(combat_dict)
        combat_dict.pop("_id", None)
        return combat_dict
    
    # ============ PREMIER TOUR ============
    first_tour = tour_names[0]
    num_first_tour_combats = bracket_size // 2
    
    # Pour chaque combat du premier tour
    for i in range(num_first_tour_combats):
        slot1 = slots[i * 2]
        slot2 = slots[i * 2 + 1]
        
        p1_id = slot1["competiteur_id"] if slot1 and isinstance(slot1, dict) else None
        p2_id = slot2["competiteur_id"] if slot2 and isinstance(slot2, dict) else None
        
        position = i + 1
        
        # Cas 1: Deux combattants réels → combat normal
        if p1_id and p2_id:
            combat_dict = await insert_combat(
                competition_id, categorie_id, tatami_id,
                first_tour, position, p1_id, p2_id, pret=True
            )
            combats_created.append(combat_dict)
            combats_by_tour[first_tour].append(combat_dict)
            combat_map[first_tour][position] = {"combat_id": combat_dict["combat_id"], "bye_winner": None}
        
        # Cas 2: Un seul combattant (BYE) → le combattant passe au tour suivant
        elif p1_id and not p2_id:
            # p1 a un BYE, on ne crée pas de combat mais on note le passage
            combat_map[first_tour][position] = {"combat_id": None, "bye_winner": p1_id}
        
        elif p2_id and not p1_id:
            # p2 a un BYE, on ne crée pas de combat mais on note le passage
            combat_map[first_tour][position] = {"combat_id": None, "bye_winner": p2_id}
        
        # Cas 3: Deux BYEs → ne devrait pas arriver avec une bonne distribution
        else:
            combat_map[first_tour][position] = {"combat_id": None, "bye_winner": None}
    
    # ============ TOURS SUIVANTS (tous obligatoires) ============
    for tour_idx in range(1, len(tour_names)):
        current_tour = tour_names[tour_idx]
        prev_tour = tour_names[tour_idx - 1]
        
        num_combats = bracket_size // (2 ** (tour_idx + 1))
        
        for i in range(num_combats):
            position = i + 1
            
            # Les 2 combats/slots sources du tour précédent
            source_pos_1 = i * 2 + 1
            source_pos_2 = i * 2 + 2
            
            source_1 = combat_map[prev_tour].get(source_pos_1, {})
            source_2 = combat_map[prev_tour].get(source_pos_2, {})
            
            # Déterminer qui vient de chaque source
            # Si c'est un BYE, le gagnant est déjà connu
            # Si c'est un combat réel, le gagnant est "À déterminer" (None)
            rouge_id = source_1.get("bye_winner")  # None si combat réel
            bleu_id = source_2.get("bye_winner")   # None si combat réel
            
            # Un combat est prêt si les deux combattants sont connus
            pret = (rouge_id is not None) and (bleu_id is not None)
            
            # Créer le combat
            combat_dict = await insert_combat(
                competition_id, categorie_id, tatami_id,
                current_tour, position, rouge_id, bleu_id, pret=pret
            )
            combats_created.append(combat_dict)
            combats_by_tour[current_tour].append(combat_dict)
            combat_map[current_tour][position] = {"combat_id": combat_dict["combat_id"], "bye_winner": None}
    
    # ============ MISE À JOUR DES LIENS DE PROPAGATION ============
    # Mettre à jour chaque combat avec le combat_suivant_id et combat_suivant_slot
    for tour_idx in range(len(tour_names) - 1):
        current_tour = tour_names[tour_idx]
        next_tour = tour_names[tour_idx + 1]
        
        combats_current = combats_by_tour[current_tour]
        
        for combat in combats_current:
            position = combat["position"]
            
            # Le combat suivant est au tour+1, position = (position+1)//2
            next_position = (position + 1) // 2
            next_slot = "rouge" if position % 2 == 1 else "bleu"
            
            # Trouver le combat suivant
            next_combat = next((c for c in combats_by_tour[next_tour] if c["position"] == next_position), None)
            
            if next_combat:
                await db.combats.update_one(
                    {"combat_id": combat["combat_id"]},
                    {"$set": {
                        "combat_suivant_id": next_combat["combat_id"],
                        "combat_suivant_slot": next_slot
                    }}
                )
    
    # ============ VÉRIFICATIONS ============
    # Nombre total de combats réels = n - 1 (un éliminé par combat jusqu'au champion)
    expected_real_combats = n - 1
    actual_real_combats = len([c for c in combats_created if c.get("rouge_id") or c.get("bleu_id")])
    
    # Nombre de combats par tour
    for tour in tour_names:
        expected_count = bracket_size // (2 ** (tour_names.index(tour) + 1))
        # Note: certains combats du premier tour ne sont pas créés (BYEs purs)
    
    # Mettre à jour la catégorie
    await db.categories.update_one(
        {"categorie_id": categorie_id},
        {"$set": {
            "nb_combattants": n, 
            "arbre_genere": True,
            "bracket_size": bracket_size,
            "num_byes": num_byes,
            "byes_locked": False
        }}
    )
    
    return {
        "message": f"Tableau généré avec {len(combats_created)} combats pour {n} combattants (bracket de {bracket_size})",
        "combats": combats_created,
        "nb_combattants": n,
        "bracket_size": bracket_size,
        "nb_byes": num_byes,
        "tours": tour_names,
        "combats_par_tour": {tour: len(combats_by_tour[tour]) for tour in tour_names}
    }

@api_router.put("/combats/{combat_id}/resultat")
async def saisir_resultat(combat_id: str, data: CombatResultat, user: User = Depends(require_admin)):
    """Saisir ou modifier le résultat d'un combat"""
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    # Sauvegarder dans l'historique si modification
    if combat.get("termine"):
        historique = HistoriqueResultat(
            combat_id=combat_id,
            ancien_vainqueur_id=combat.get("vainqueur_id"),
            nouveau_vainqueur_id=data.vainqueur_id,
            ancien_score_rouge=combat.get("score_rouge", 0),
            ancien_score_bleu=combat.get("score_bleu", 0),
            nouveau_score_rouge=data.score_rouge,
            nouveau_score_bleu=data.score_bleu,
            modifie_par=user.user_id
        )
        hist_dict = historique.model_dump()
        hist_dict["modifie_at"] = hist_dict["modifie_at"].isoformat()
        await db.historique_resultats.insert_one(hist_dict)
    
    # Gérer la disqualification
    if data.type_victoire == "disqualification":
        perdant_id = combat["rouge_id"] if data.vainqueur_id == combat["bleu_id"] else combat["bleu_id"]
        await db.competiteurs.update_one(
            {"competiteur_id": perdant_id},
            {"$set": {"disqualifie": True}}
        )
    
    # Mettre à jour le combat
    await db.combats.update_one(
        {"combat_id": combat_id},
        {"$set": {
            "vainqueur_id": data.vainqueur_id,
            "score_rouge": data.score_rouge,
            "score_bleu": data.score_bleu,
            "type_victoire": data.type_victoire,
            "termine": True,
            "statut": "termine"
        }}
    )
    
    # Propager le vainqueur au tour suivant
    await propager_vainqueur(combat, data.vainqueur_id)
    
    updated = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    return updated

async def propager_vainqueur(combat: dict, vainqueur_id: str):
    """
    Propage le vainqueur au combat suivant.
    Supporte à la fois:
    - Le système automatique (basé sur tour/position)
    - Le système manuel (basé sur combat_suivant_id/combat_suivant_slot)
    
    RÈGLES TAEKWONDO:
    - PAS de combat pour la 3ème place
    - Les perdants des demi-finales reçoivent le bronze automatiquement
    """
    categorie_id = combat["categorie_id"]
    tour = combat["tour"]
    position = combat["position"]
    
    # ========== PROPAGATION MANUELLE ==========
    # Si le combat a une connexion manuelle définie, l'utiliser en priorité
    if combat.get("combat_suivant_id") and combat.get("combat_suivant_slot"):
        field = "rouge_id" if combat["combat_suivant_slot"] == "rouge" else "bleu_id"
        await db.combats.update_one(
            {"combat_id": combat["combat_suivant_id"]},
            {"$set": {field: vainqueur_id}}
        )
        # Mettre à jour le statut prêt du combat suivant
        await update_combat_pret_status(combat["combat_suivant_id"])
        return  # Terminé pour les connexions manuelles
    
    # ========== PROPAGATION AUTOMATIQUE (système existant) ==========
    if tour == "seizieme":
        # Vers huitième de finale
        huitieme_position = (position + 1) // 2
        huitieme = await db.combats.find_one({
            "categorie_id": categorie_id,
            "tour": "huitieme",
            "position": huitieme_position
        }, {"_id": 0})
        
        if huitieme:
            field = "rouge_id" if position % 2 == 1 else "bleu_id"
            await db.combats.update_one(
                {"combat_id": huitieme["combat_id"]},
                {"$set": {field: vainqueur_id}}
            )
            await update_combat_pret_status(huitieme["combat_id"])
    
    elif tour == "huitieme":
        # Vers quart de finale
        quart_position = (position + 1) // 2
        quart = await db.combats.find_one({
            "categorie_id": categorie_id,
            "tour": "quart",
            "position": quart_position
        }, {"_id": 0})
        
        if quart:
            field = "rouge_id" if position % 2 == 1 else "bleu_id"
            await db.combats.update_one(
                {"combat_id": quart["combat_id"]},
                {"$set": {field: vainqueur_id}}
            )
            await update_combat_pret_status(quart["combat_id"])
    
    elif tour == "quart":
        # Vers demi-finale
        demi_position = (position + 1) // 2
        demi = await db.combats.find_one({
            "categorie_id": categorie_id,
            "tour": "demi",
            "position": demi_position
        }, {"_id": 0})
        
        if demi:
            field = "rouge_id" if position % 2 == 1 else "bleu_id"
            await db.combats.update_one(
                {"combat_id": demi["combat_id"]},
                {"$set": {field: vainqueur_id}}
            )
            await update_combat_pret_status(demi["combat_id"])
    
    elif tour == "demi":
        # Vers finale (vainqueur uniquement)
        # PAS de match bronze en Taekwondo - les perdants des demis ont le bronze ex-aequo
        finale = await db.combats.find_one({
            "categorie_id": categorie_id,
            "tour": "finale"
        }, {"_id": 0})
        
        if finale:
            field = "rouge_id" if position == 1 else "bleu_id"
            await db.combats.update_one(
                {"combat_id": finale["combat_id"]},
                {"$set": {field: vainqueur_id}}
            )
            await update_combat_pret_status(finale["combat_id"])

@api_router.post("/combats/{categorie_id}/attribuer-medailles")
async def attribuer_medailles(categorie_id: str, user: User = Depends(require_admin)):
    """
    Attribue les médailles après la finale selon les règles World Taekwondo.
    
    PODIUM:
    - 🥇 Or: Vainqueur de la finale
    - 🥈 Argent: Perdant de la finale
    - 🥉 Bronze: Perdant demi-finale 1 (ex-aequo)
    - 🥉 Bronze: Perdant demi-finale 2 (ex-aequo)
    
    PAS de combat pour la 3ème place en Taekwondo.
    """
    # Finale
    finale = await db.combats.find_one({
        "categorie_id": categorie_id,
        "tour": "finale",
        "termine": True
    }, {"_id": 0})
    
    if not finale:
        raise HTTPException(status_code=400, detail="La finale n'est pas terminée")
    
    # Supprimer les anciennes médailles
    await db.medailles.delete_many({"categorie_id": categorie_id})
    
    medailles = []
    
    # Or au vainqueur de la finale
    or_medaille = Medaille(
        categorie_id=categorie_id,
        competiteur_id=finale["vainqueur_id"],
        type="or"
    )
    await db.medailles.insert_one(or_medaille.model_dump())
    medailles.append(or_medaille.model_dump())
    
    # Argent au perdant de la finale (si non disqualifié)
    perdant_finale = finale["rouge_id"] if finale["vainqueur_id"] == finale["bleu_id"] else finale["bleu_id"]
    if perdant_finale:
        perdant = await db.competiteurs.find_one({"competiteur_id": perdant_finale}, {"_id": 0})
        if perdant and not perdant.get("disqualifie"):
            argent_medaille = Medaille(
                categorie_id=categorie_id,
                competiteur_id=perdant_finale,
                type="argent"
            )
            await db.medailles.insert_one(argent_medaille.model_dump())
            medailles.append(argent_medaille.model_dump())
    
    # Bronze ex-aequo aux perdants des demi-finales (PAS de combat bronze en Taekwondo)
    demis = await db.combats.find({
        "categorie_id": categorie_id,
        "tour": "demi",
        "termine": True
    }, {"_id": 0}).to_list(10)
    
    for demi in demis:
        if demi.get("vainqueur_id"):
            perdant_id = demi["rouge_id"] if demi["vainqueur_id"] == demi["bleu_id"] else demi["bleu_id"]
            if perdant_id:
                perdant = await db.competiteurs.find_one({"competiteur_id": perdant_id}, {"_id": 0})
                if perdant and not perdant.get("disqualifie"):
                    bronze_medaille = Medaille(
                        categorie_id=categorie_id,
                        competiteur_id=perdant_id,
                        type="bronze"
                    )
                    await db.medailles.insert_one(bronze_medaille.model_dump())
                    medailles.append(bronze_medaille.model_dump())
    
    return {"message": f"{len(medailles)} médailles attribuées", "medailles": medailles}

# ============ GESTION DES BYES ============

class ByeModificationRequest(BaseModel):
    """Requête pour modifier les BYEs d'une catégorie"""
    competiteur_ids_with_bye: List[str]  # IDs des compétiteurs qui auront un BYE

@api_router.get("/categories/{categorie_id}/byes")
async def get_byes_info(categorie_id: str, user: User = Depends(get_current_user)):
    """
    Récupère les informations sur les BYEs d'une catégorie.
    Inclut la liste des compétiteurs et leur statut BYE.
    """
    categorie = await db.categories.find_one({"categorie_id": categorie_id}, {"_id": 0})
    if not categorie:
        raise HTTPException(status_code=404, detail="Catégorie non trouvée")
    
    # Récupérer tous les compétiteurs de la catégorie
    competiteurs = await db.competiteurs.find(
        {"categorie_id": categorie_id, "disqualifie": False},
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    
    n = len(competiteurs)
    
    # Calculer la taille du bracket
    def next_power_of_2(x):
        if x <= 1:
            return 1
        return 1 << (x - 1).bit_length()
    
    bracket_size = next_power_of_2(n) if n > 0 else 0
    num_byes = bracket_size - n if n > 0 else 0
    
    # Vérifier si les BYEs sont verrouillés (un combat a déjà commencé)
    combat_en_cours = await db.combats.find_one({
        "categorie_id": categorie_id,
        "statut": {"$in": ["en_cours", "termine"]}
    })
    byes_locked = combat_en_cours is not None
    
    # Récupérer les compétiteurs actuellement avec BYE
    # Un compétiteur a un BYE si son ID apparaît seul dans un combat du premier tour
    premier_tour = "seizieme" if bracket_size == 32 else "huitieme" if bracket_size == 16 else "quart" if bracket_size == 8 else "demi" if bracket_size == 4 else "finale"
    
    combats_premier_tour = await db.combats.find({
        "categorie_id": categorie_id,
        "tour": premier_tour
    }, {"_id": 0}).to_list(50)
    
    # Identifier les compétiteurs avec BYE (ceux qui ne sont pas dans les combats du premier tour)
    competiteurs_en_combat = set()
    for combat in combats_premier_tour:
        if combat.get("rouge_id"):
            competiteurs_en_combat.add(combat["rouge_id"])
        if combat.get("bleu_id"):
            competiteurs_en_combat.add(combat["bleu_id"])
    
    competiteurs_avec_bye = []
    for c in competiteurs:
        if c["competiteur_id"] not in competiteurs_en_combat:
            competiteurs_avec_bye.append(c["competiteur_id"])
    
    return {
        "categorie_id": categorie_id,
        "categorie_nom": categorie.get("nom", ""),
        "nb_competiteurs": n,
        "bracket_size": bracket_size,
        "num_byes_disponibles": num_byes,
        "byes_locked": byes_locked,
        "competiteurs_avec_bye": competiteurs_avec_bye,
        "competiteurs": [
            {
                "competiteur_id": c["competiteur_id"],
                "nom": c.get("nom", ""),
                "prenom": c.get("prenom", ""),
                "club": c.get("club", ""),
                "has_bye": c["competiteur_id"] in competiteurs_avec_bye
            }
            for c in competiteurs
        ]
    }

@api_router.put("/categories/{categorie_id}/byes")
async def modifier_byes(categorie_id: str, data: ByeModificationRequest, user: User = Depends(require_admin)):
    """
    Modifie l'attribution des BYEs pour une catégorie.
    Réservé au MASTER/Admin.
    Ne fonctionne que si aucun combat n'a encore commencé.
    """
    # Vérifier que l'utilisateur est MASTER
    if user.role != "master":
        raise HTTPException(status_code=403, detail="Seul le MASTER peut modifier les BYEs")
    
    categorie = await db.categories.find_one({"categorie_id": categorie_id}, {"_id": 0})
    if not categorie:
        raise HTTPException(status_code=404, detail="Catégorie non trouvée")
    
    # Vérifier si les BYEs sont verrouillés
    combat_en_cours = await db.combats.find_one({
        "categorie_id": categorie_id,
        "statut": {"$in": ["en_cours", "termine"]}
    })
    if combat_en_cours:
        raise HTTPException(
            status_code=400, 
            detail="🔒 Les BYEs sont verrouillés car un combat a déjà commencé dans cette catégorie"
        )
    
    # Récupérer les compétiteurs
    competiteurs = await db.competiteurs.find(
        {"categorie_id": categorie_id, "disqualifie": False},
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    
    n = len(competiteurs)
    
    # Calculer le nombre de BYEs disponibles
    def next_power_of_2(x):
        if x <= 1:
            return 1
        return 1 << (x - 1).bit_length()
    
    bracket_size = next_power_of_2(n)
    num_byes = bracket_size - n
    
    # Vérifier que le nombre de BYEs demandé est correct
    if len(data.competiteur_ids_with_bye) != num_byes:
        raise HTTPException(
            status_code=400,
            detail=f"Vous devez attribuer exactement {num_byes} BYE(s) pour {n} compétiteurs dans un bracket de {bracket_size}"
        )
    
    # Vérifier que tous les IDs sont valides
    competiteur_ids = {c["competiteur_id"] for c in competiteurs}
    for cid in data.competiteur_ids_with_bye:
        if cid not in competiteur_ids:
            raise HTTPException(status_code=400, detail=f"Compétiteur {cid} non trouvé dans cette catégorie")
    
    # Régénérer l'arbre avec les nouveaux BYEs
    competition_id = categorie.get("competition_id")
    
    # Supprimer les anciens combats
    await db.combats.delete_many({"categorie_id": categorie_id})
    
    # Créer la liste des compétiteurs avec les BYEs en premier
    bye_set = set(data.competiteur_ids_with_bye)
    competiteurs_sans_bye = [c for c in competiteurs if c["competiteur_id"] not in bye_set]
    competiteurs_avec_bye = [c for c in competiteurs if c["competiteur_id"] in bye_set]
    
    # Mélanger les deux groupes séparément pour le tirage au sort
    random.shuffle(competiteurs_sans_bye)
    random.shuffle(competiteurs_avec_bye)
    
    # Les combattants avec BYE seront placés stratégiquement
    combats_created = []
    
    async def insert_combat(combat_obj):
        combat_dict = combat_obj.model_dump()
        combat_dict["created_at"] = combat_dict["created_at"].isoformat()
        await db.combats.insert_one(combat_dict)
        combat_dict.pop("_id", None)
        return combat_dict
    
    def get_tour_name(num_participants_at_round):
        if num_participants_at_round == 2:
            return "finale"
        elif num_participants_at_round == 4:
            return "demi"
        elif num_participants_at_round == 8:
            return "quart"
        elif num_participants_at_round == 16:
            return "huitieme"
        elif num_participants_at_round == 32:
            return "seizieme"
        else:
            return f"tour_{num_participants_at_round}"
    
    # Créer les slots avec BYEs aux bonnes positions
    slots = [None] * bracket_size
    
    # Placer les BYEs de manière stratégique (positions impaires pour répartition)
    bye_positions = []
    if num_byes > 0:
        step = bracket_size // num_byes if num_byes > 0 else bracket_size
        for i in range(num_byes):
            pos = (i * step + 1) % bracket_size
            while pos in bye_positions and len(bye_positions) < num_byes:
                pos = (pos + 1) % bracket_size
            bye_positions.append(pos)
    
    # Placer les compétiteurs avec BYE en face des positions de BYE
    bye_idx = 0
    sans_bye_idx = 0
    
    for i in range(bracket_size):
        if i in bye_positions:
            slots[i] = None  # BYE
        else:
            # Placer un compétiteur avec BYE s'il y en a encore et que c'est une position adjacente à un BYE
            adjacent_is_bye = (i > 0 and (i-1) in bye_positions) or (i < bracket_size-1 and (i+1) in bye_positions)
            
            if adjacent_is_bye and bye_idx < len(competiteurs_avec_bye):
                slots[i] = competiteurs_avec_bye[bye_idx]
                bye_idx += 1
            elif sans_bye_idx < len(competiteurs_sans_bye):
                slots[i] = competiteurs_sans_bye[sans_bye_idx]
                sans_bye_idx += 1
            elif bye_idx < len(competiteurs_avec_bye):
                slots[i] = competiteurs_avec_bye[bye_idx]
                bye_idx += 1
    
    # Générer les combats
    current_slots = slots.copy()
    tour_size = bracket_size
    
    while tour_size >= 2:
        tour_name = get_tour_name(tour_size)
        num_combats = tour_size // 2
        next_slots = []
        
        for i in range(num_combats):
            p1 = current_slots[i * 2] if i * 2 < len(current_slots) else None
            p2 = current_slots[i * 2 + 1] if i * 2 + 1 < len(current_slots) else None
            
            p1_id = p1["competiteur_id"] if p1 and isinstance(p1, dict) and "competiteur_id" in p1 else (p1 if isinstance(p1, str) else None)
            p2_id = p2["competiteur_id"] if p2 and isinstance(p2, dict) and "competiteur_id" in p2 else (p2 if isinstance(p2, str) else None)
            
            if p1_id is None and p2_id is not None:
                next_slots.append(p2_id)
                continue
            elif p2_id is None and p1_id is not None:
                next_slots.append(p1_id)
                continue
            elif p1_id is None and p2_id is None:
                next_slots.append(None)
                continue
            
            combat = Combat(
                competition_id=competition_id,
                categorie_id=categorie_id,
                tour=tour_name,
                position=i + 1,
                rouge_id=p1_id,
                bleu_id=p2_id,
                has_bye=False,
                pret=True
            )
            combat_dict = await insert_combat(combat)
            combats_created.append(combat_dict)
            next_slots.append({"placeholder": combat_dict["combat_id"]})
        
        if tour_size > 2:
            current_slots = next_slots
        tour_size = tour_size // 2
    
    # Créer les combats des tours suivants s'ils n'existent pas
    if bracket_size >= 4:
        demi_exists = await db.combats.find_one({"categorie_id": categorie_id, "tour": "demi"})
        if not demi_exists:
            for i in range(2):
                demi = Combat(
                    competition_id=competition_id,
                    categorie_id=categorie_id,
                    tour="demi",
                    position=i + 1,
                    rouge_id=None,
                    bleu_id=None,
                    pret=False
                )
                demi_dict = await insert_combat(demi)
                combats_created.append(demi_dict)
    
    finale_exists = await db.combats.find_one({"categorie_id": categorie_id, "tour": "finale"})
    if not finale_exists:
        finale = Combat(
            competition_id=competition_id,
            categorie_id=categorie_id,
            tour="finale",
            position=1,
            rouge_id=None,
            bleu_id=None,
            est_finale=True,
            pret=False
        )
        finale_dict = await insert_combat(finale)
        combats_created.append(finale_dict)
    
    return {
        "message": f"BYEs modifiés avec succès. {len(combats_created)} combats régénérés.",
        "competiteurs_avec_bye": data.competiteur_ids_with_bye,
        "combats": combats_created
    }

@api_router.put("/combats/{combat_id}/statut")
async def modifier_statut_combat(combat_id: str, statut: str, user: User = Depends(require_admin)):
    """Modifier le statut d'un combat (a_venir, en_cours, termine, non_dispute)"""
    if statut not in ["a_venir", "en_cours", "termine", "non_dispute"]:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    result = await db.combats.update_one(
        {"combat_id": combat_id},
        {"$set": {"statut": statut, "termine": statut in ["termine", "non_dispute"]}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    return {"message": "Statut mis à jour"}

@api_router.post("/combats/planifier/{categorie_id}")
async def planifier_combats(categorie_id: str, data: PlanificationCreate, user: User = Depends(require_admin)):
    """Planifier les horaires des combats d'une catégorie"""
    combats = await db.combats.find(
        {"categorie_id": categorie_id},
        {"_id": 0}
    ).sort([("tour", 1), ("position", 1)]).to_list(100)
    
    if not combats:
        raise HTTPException(status_code=404, detail="Aucun combat trouvé")
    
    # Définir l'ordre des tours
    tour_ordre = {"quart": 1, "demi": 2, "bronze": 3, "finale": 4}
    combats_sorted = sorted(combats, key=lambda x: (tour_ordre.get(x["tour"], 99), x["position"]))
    
    # Parser l'heure de début
    try:
        heure_parts = data.heure_debut_competition.split(":")
        current_hour = int(heure_parts[0])
        current_minute = int(heure_parts[1])
    except:
        raise HTTPException(status_code=400, detail="Format d'heure invalide (HH:MM)")
    
    # Planifier chaque combat
    pauses_dict = {p.get("apres_combat", 0): p.get("duree_minutes", 15) for p in data.pauses}
    
    for i, combat in enumerate(combats_sorted):
        heure = f"{current_hour:02d}:{current_minute:02d}"
        
        await db.combats.update_one(
            {"combat_id": combat["combat_id"]},
            {"$set": {
                "ordre": i + 1,
                "heure_debut": heure,
                "duree_minutes": data.duree_combat_minutes
            }}
        )
        
        # Ajouter la durée du combat
        current_minute += data.duree_combat_minutes
        
        # Ajouter une pause si définie
        if (i + 1) in pauses_dict:
            current_minute += pauses_dict[i + 1]
        
        # Gérer le dépassement d'heure
        while current_minute >= 60:
            current_minute -= 60
            current_hour += 1
    
    return {"message": f"{len(combats_sorted)} combats planifiés", "heure_fin_estimee": f"{current_hour:02d}:{current_minute:02d}"}

@api_router.put("/combats/modifier-ordre")
async def modifier_ordre_combat(data: ModifierOrdreCombat, user: User = Depends(require_admin)):
    """Modifier l'ordre d'un combat et recalculer les horaires si nécessaire"""
    combat = await db.combats.find_one({"combat_id": data.combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    update_data = {"ordre": data.nouvel_ordre}
    if data.nouvelle_heure:
        update_data["heure_debut"] = data.nouvelle_heure
    
    await db.combats.update_one(
        {"combat_id": data.combat_id},
        {"$set": update_data}
    )
    
    return {"message": "Ordre mis à jour"}

@api_router.post("/combats/lancer-categorie/{categorie_id}")
async def lancer_categorie(categorie_id: str, mode: str = "complet", user: User = Depends(require_admin)):
    """
    Lancer les combats d'une catégorie
    mode: 'complet' = tous les combats, 'finales_fin' = tous sauf finales
    """
    if mode not in ["complet", "finales_fin"]:
        raise HTTPException(status_code=400, detail="Mode invalide (complet ou finales_fin)")
    
    query = {"categorie_id": categorie_id, "statut": "a_venir"}
    if mode == "finales_fin":
        query["tour"] = {"$nin": ["finale", "bronze"]}
    
    # Mettre le premier combat en cours
    premier_combat = await db.combats.find_one(query, {"_id": 0}, sort=[("ordre", 1)])
    
    if premier_combat:
        await db.combats.update_one(
            {"combat_id": premier_combat["combat_id"]},
            {"$set": {"statut": "en_cours"}}
        )
        return {"message": f"Catégorie lancée en mode {mode}", "premier_combat": premier_combat["combat_id"]}
    
    return {"message": "Aucun combat à lancer"}

@api_router.post("/combats/lancer-finales")
async def lancer_finales(user: User = Depends(require_admin)):
    """Lancer toutes les finales de toutes les catégories"""
    finales = await db.combats.find(
        {"tour": {"$in": ["finale", "bronze"]}, "statut": "a_venir"},
        {"_id": 0}
    ).sort("ordre", 1).to_list(100)
    
    if finales:
        # Mettre la première finale en cours
        await db.combats.update_one(
            {"combat_id": finales[0]["combat_id"]},
            {"$set": {"statut": "en_cours"}}
        )
        return {"message": f"{len(finales)} finales prêtes à être lancées", "premiere_finale": finales[0]["combat_id"]}
    
    return {"message": "Aucune finale à lancer"}

@api_router.post("/combats/{combat_id}/suivant")
async def passer_combat_suivant(combat_id: str, user: User = Depends(require_admin)):
    """Terminer le combat actuel et passer au suivant"""
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    # Trouver le prochain combat
    prochain = await db.combats.find_one(
        {"ordre": {"$gt": combat.get("ordre", 0)}, "statut": "a_venir"},
        {"_id": 0},
        sort=[("ordre", 1)]
    )
    
    if prochain:
        await db.combats.update_one(
            {"combat_id": prochain["combat_id"]},
            {"$set": {"statut": "en_cours"}}
        )
        return {"message": "Combat suivant lancé", "combat_id": prochain["combat_id"]}
    
    return {"message": "Plus de combat à suivre"}

# ============ MEDAILLES ENDPOINTS ============

@api_router.get("/medailles", response_model=List[Medaille])
async def list_medailles(categorie_id: Optional[str] = None, user: User = Depends(get_current_user)):
    query = {}
    if categorie_id:
        query["categorie_id"] = categorie_id
    
    medailles = await db.medailles.find(query, {"_id": 0}).to_list(1000)
    return medailles

# ============ HISTORIQUE ENDPOINTS ============

@api_router.get("/historique", response_model=List[HistoriqueResultat])
async def list_historique(combat_id: Optional[str] = None, user: User = Depends(get_current_user)):
    query = {}
    if combat_id:
        query["combat_id"] = combat_id
    
    historique = await db.historique_resultats.find(query, {"_id": 0}).to_list(1000)
    return historique

# ============ STATS ENDPOINTS ============

@api_router.get("/stats")
async def get_stats(user: User = Depends(get_current_user)):
    competiteurs_count = await db.competiteurs.count_documents({})
    categories_count = await db.categories.count_documents({})
    combats_count = await db.combats.count_documents({})
    combats_termines = await db.combats.count_documents({"termine": True})
    medailles_count = await db.medailles.count_documents({})
    tatamis_count = await db.tatamis.count_documents({})
    
    return {
        "competiteurs": competiteurs_count,
        "categories": categories_count,
        "combats_total": combats_count,
        "combats_termines": combats_termines,
        "medailles": medailles_count,
        "tatamis": tatamis_count
    }

# ============ ADMIN: Promote user ============

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: str, admin: User = Depends(require_admin)):
    """Change le rôle d'un utilisateur. Seul un MASTER peut promouvoir en master."""
    if role not in ["coach", "admin", "master"]:
        raise HTTPException(status_code=400, detail="Rôle invalide")
    
    # Seul un master peut promouvoir quelqu'un en master
    if role == "master" and admin.role != "master":
        raise HTTPException(status_code=403, detail="Seul un MASTER peut promouvoir en MASTER")
    
    # Empêcher de rétrograder un master si on n'est pas master
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "role": 1})
    if target_user and target_user.get("role") == "master" and admin.role != "master":
        raise HTTPException(status_code=403, detail="Impossible de modifier le rôle d'un MASTER")
    
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": role}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    return {"message": f"Rôle mis à jour en {role}"}

@api_router.get("/users")
async def list_users(admin: User = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    return users

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: User = Depends(require_master)):
    """Supprime un utilisateur. Réservé au MASTER."""
    if user_id == admin.user_id:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas vous supprimer vous-même")
    
    result = await db.users.delete_one({"user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Supprimer aussi les sessions de cet utilisateur
    await db.user_sessions.delete_many({"user_id": user_id})
    
    return {"message": "Utilisateur supprimé"}

# ============ ARBITRE ENDPOINTS ============

@api_router.get("/arbitre/aire/{aire_id}")
async def get_arbitre_view(aire_id: str, user: User = Depends(get_current_user)):
    """
    Vue complète pour l'arbitre de table centrale d'une aire de combat.
    Retourne le combat en cours, les combats à venir et les infos des compétiteurs.
    """
    aire = await db.aires_combat.find_one({"aire_id": aire_id}, {"_id": 0})
    if not aire:
        raise HTTPException(status_code=404, detail="Aire de combat non trouvée")
    
    # Combat en cours
    combat_en_cours = await db.combats.find_one(
        {"aire_id": aire_id, "statut": "en_cours"},
        {"_id": 0}
    )
    
    # Enrichir avec les infos des compétiteurs
    if combat_en_cours:
        if combat_en_cours.get("rouge_id"):
            rouge = await db.competiteurs.find_one(
                {"competiteur_id": combat_en_cours["rouge_id"]},
                {"_id": 0}
            )
            combat_en_cours["rouge"] = rouge
        if combat_en_cours.get("bleu_id"):
            bleu = await db.competiteurs.find_one(
                {"competiteur_id": combat_en_cours["bleu_id"]},
                {"_id": 0}
            )
            combat_en_cours["bleu"] = bleu
        # Catégorie
        categorie = await db.categories.find_one(
            {"categorie_id": combat_en_cours["categorie_id"]},
            {"_id": 0}
        )
        combat_en_cours["categorie"] = categorie
    
    # Combats à venir sur cette aire (non-finales d'abord)
    combats_a_venir = await db.combats.find(
        {"aire_id": aire_id, "statut": "a_venir"},
        {"_id": 0}
    ).sort([("est_finale", 1), ("ordre", 1)]).to_list(20)
    
    # Enrichir chaque combat à venir
    for combat in combats_a_venir:
        if combat.get("rouge_id"):
            rouge = await db.competiteurs.find_one(
                {"competiteur_id": combat["rouge_id"]},
                {"_id": 0, "competiteur_id": 1, "nom": 1, "prenom": 1, "club": 1}
            )
            combat["rouge"] = rouge
        if combat.get("bleu_id"):
            bleu = await db.competiteurs.find_one(
                {"competiteur_id": combat["bleu_id"]},
                {"_id": 0, "competiteur_id": 1, "nom": 1, "prenom": 1, "club": 1}
            )
            combat["bleu"] = bleu
        categorie = await db.categories.find_one(
            {"categorie_id": combat["categorie_id"]},
            {"_id": 0, "nom": 1}
        )
        combat["categorie"] = categorie
    
    # Finales en attente (toutes les aires confondues pour info)
    finales_restantes = await db.combats.count_documents({
        "competition_id": aire["competition_id"],
        "est_finale": True,
        "termine": False
    })
    
    return {
        "aire": aire,
        "combat_en_cours": combat_en_cours,
        "combats_a_venir": combats_a_venir,
        "finales_restantes": finales_restantes
    }

@api_router.post("/arbitre/lancer/{combat_id}")
async def lancer_combat(combat_id: str, user: User = Depends(get_current_user)):
    """Lance un combat (le passe en statut 'en_cours')"""
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    if combat["statut"] != "a_venir":
        raise HTTPException(status_code=400, detail="Ce combat ne peut pas être lancé")
    
    # Vérifier que les deux combattants sont présents
    if not combat.get("rouge_id") or not combat.get("bleu_id"):
        raise HTTPException(status_code=400, detail="Les deux combattants doivent être définis")
    
    # Passer le combat en cours
    await db.combats.update_one(
        {"combat_id": combat_id},
        {"$set": {"statut": "en_cours"}}
    )
    
    updated = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    return updated

@api_router.post("/arbitre/resultat/{combat_id}")
async def saisir_resultat_rapide(
    combat_id: str, 
    vainqueur: str,  # "rouge" ou "bleu"
    score_rouge: int = 0,
    score_bleu: int = 0,
    type_victoire: str = "normal",
    user: User = Depends(get_current_user)
):
    """
    Saisie rapide du résultat par l'arbitre.
    Le perdant est automatiquement marqué comme éliminé.
    """
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    if combat["statut"] != "en_cours":
        raise HTTPException(status_code=400, detail="Ce combat n'est pas en cours")
    
    # Déterminer le vainqueur
    if vainqueur == "rouge":
        vainqueur_id = combat["rouge_id"]
        perdant_id = combat["bleu_id"]
    elif vainqueur == "bleu":
        vainqueur_id = combat["bleu_id"]
        perdant_id = combat["rouge_id"]
    else:
        raise HTTPException(status_code=400, detail="Le vainqueur doit être 'rouge' ou 'bleu'")
    
    # En Taekwondo, le perdant est éliminé (ne peut plus combattre)
    # Sauf s'il y a un match pour le bronze
    if perdant_id:
        # Vérifier si ce n'est pas une demi-finale (où le perdant peut aller au bronze)
        is_demi = combat["tour"] == "demi"
        if not is_demi:
            # Marquer comme éliminé définitivement
            await db.competiteurs.update_one(
                {"competiteur_id": perdant_id},
                {"$set": {"elimine": True}}
            )
    
    # Mettre à jour le combat
    await db.combats.update_one(
        {"combat_id": combat_id},
        {"$set": {
            "vainqueur_id": vainqueur_id,
            "score_rouge": score_rouge,
            "score_bleu": score_bleu,
            "type_victoire": type_victoire,
            "termine": True,
            "statut": "termine"
        }}
    )
    
    # Propager le vainqueur au tour suivant
    await propager_vainqueur(combat, vainqueur_id)
    
    # Propager le perdant au match bronze si c'est une demi-finale
    if combat["tour"] == "demi" and perdant_id:
        bronze_match = await db.combats.find_one({
            "categorie_id": combat["categorie_id"],
            "tour": "bronze"
        }, {"_id": 0})
        
        if bronze_match:
            # Déterminer quelle position (rouge ou bleu) selon la position de la demi
            update_field = "rouge_id" if combat["position"] == 1 else "bleu_id"
            await db.combats.update_one(
                {"combat_id": bronze_match["combat_id"]},
                {"$set": {update_field: perdant_id}}
            )
    
    updated = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    return updated

@api_router.get("/arbitre/prochain/{aire_id}")
async def get_prochain_combat(aire_id: str, user: User = Depends(get_current_user)):
    """Récupère le prochain combat à lancer sur une aire"""
    # D'abord les combats non-finales
    combat = await db.combats.find_one(
        {"aire_id": aire_id, "statut": "a_venir", "est_finale": False},
        {"_id": 0},
        sort=[("ordre", 1)]
    )
    
    # Si pas de combat régulier, chercher les finales
    if not combat:
        combat = await db.combats.find_one(
            {"aire_id": aire_id, "statut": "a_venir", "est_finale": True},
            {"_id": 0},
            sort=[("ordre", 1)]
        )
    
    if not combat:
        return None
    
    # Enrichir avec les infos
    if combat.get("rouge_id"):
        rouge = await db.competiteurs.find_one(
            {"competiteur_id": combat["rouge_id"]},
            {"_id": 0}
        )
        combat["rouge"] = rouge
    if combat.get("bleu_id"):
        bleu = await db.competiteurs.find_one(
            {"competiteur_id": combat["bleu_id"]},
            {"_id": 0}
        )
        combat["bleu"] = bleu
    
    categorie = await db.categories.find_one(
        {"categorie_id": combat["categorie_id"]},
        {"_id": 0}
    )
    combat["categorie"] = categorie
    
    return combat

@api_router.post("/arbitre/verifier-finales/{competition_id}")
async def verifier_lancement_finales(competition_id: str, user: User = Depends(get_current_user)):
    """
    Vérifie si tous les combats non-finales sont terminés.
    Si oui, les finales peuvent commencer.
    """
    # Compter les combats non-finales non terminés
    combats_restants = await db.combats.count_documents({
        "competition_id": competition_id,
        "est_finale": False,
        "termine": False
    })
    
    # Compter les finales
    finales_total = await db.combats.count_documents({
        "competition_id": competition_id,
        "est_finale": True
    })
    
    finales_terminees = await db.combats.count_documents({
        "competition_id": competition_id,
        "est_finale": True,
        "termine": True
    })
    
    peut_lancer_finales = combats_restants == 0
    
    return {
        "combats_reguliers_restants": combats_restants,
        "finales_total": finales_total,
        "finales_terminees": finales_terminees,
        "peut_lancer_finales": peut_lancer_finales,
        "message": "Les finales peuvent commencer !" if peut_lancer_finales else f"Il reste {combats_restants} combat(s) régulier(s) à terminer"
    }

# ============ VALIDATION DES COACHS PAR COMPETITION ============

@api_router.get("/competitions/{competition_id}/coaches")
async def get_competition_coaches(competition_id: str, admin: User = Depends(require_admin)):
    """Liste les coachs autorisés pour une compétition avec leurs infos"""
    competition = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    coaches_ids = competition.get("coaches_autorises", [])
    coaches = []
    
    for coach_id in coaches_ids:
        coach = await db.users.find_one(
            {"user_id": coach_id},
            {"_id": 0, "password": 0}
        )
        if coach:
            coaches.append(coach)
    
    return coaches

@api_router.get("/competitions/{competition_id}/coaches/available")
async def get_available_coaches(competition_id: str, admin: User = Depends(require_admin)):
    """Liste les coachs NON encore autorisés pour cette compétition"""
    competition = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    coaches_autorises = competition.get("coaches_autorises", [])
    
    # Récupérer tous les coachs qui ne sont pas encore autorisés
    available_coaches = await db.users.find(
        {"role": "coach", "user_id": {"$nin": coaches_autorises}},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return available_coaches

@api_router.post("/competitions/{competition_id}/coaches/{coach_id}")
async def add_coach_to_competition(competition_id: str, coach_id: str, admin: User = Depends(require_admin)):
    """Ajoute un coach à la liste des coachs autorisés pour une compétition"""
    # Vérifier que la compétition existe
    competition = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    # Vérifier que l'utilisateur existe et est coach
    coach = await db.users.find_one({"user_id": coach_id}, {"_id": 0})
    if not coach:
        raise HTTPException(status_code=404, detail="Coach non trouvé")
    if coach.get("role") not in ["coach", "admin", "master"]:
        raise HTTPException(status_code=400, detail="L'utilisateur n'est pas un coach")
    
    # Ajouter le coach s'il n'est pas déjà autorisé
    if coach_id in competition.get("coaches_autorises", []):
        raise HTTPException(status_code=400, detail="Coach déjà autorisé pour cette compétition")
    
    await db.competitions.update_one(
        {"competition_id": competition_id},
        {"$addToSet": {"coaches_autorises": coach_id}}
    )
    
    return {"message": f"Coach {coach.get('name')} autorisé pour la compétition"}

@api_router.delete("/competitions/{competition_id}/coaches/{coach_id}")
async def remove_coach_from_competition(competition_id: str, coach_id: str, admin: User = Depends(require_admin)):
    """Retire un coach de la liste des coachs autorisés pour une compétition"""
    result = await db.competitions.update_one(
        {"competition_id": competition_id},
        {"$pull": {"coaches_autorises": coach_id}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    return {"message": "Coach retiré de la compétition"}

# ============ IMPORT/EXPORT EXCEL COMPETITEURS ============

@api_router.get("/excel/competiteurs/export/{competition_id}")
async def export_competiteurs_excel(competition_id: str, user: User = Depends(get_current_user)):
    """Exporte la liste des compétiteurs d'une compétition au format Excel"""
    if not await user_can_access_competition(user, competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    # Récupérer la compétition
    competition = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    # Récupérer les compétiteurs
    competiteurs = await db.competiteurs.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Récupérer les catégories pour les noms
    categories = await db.categories.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).to_list(500)
    cat_dict = {c["categorie_id"]: c["nom"] for c in categories}
    
    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Compétiteurs"
    
    # Style pour l'en-tête
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ["Nom", "Prénom", "Date de naissance", "Sexe", "Poids déclaré", "Poids officiel", "Club", "Catégorie", "Pesé", "Surclassé"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Données
    for row, comp in enumerate(competiteurs, 2):
        ws.cell(row=row, column=1, value=comp.get("nom", "")).border = thin_border
        ws.cell(row=row, column=2, value=comp.get("prenom", "")).border = thin_border
        # Convertir la date au format français JJ/MM/AAAA
        date_fr = date_iso_to_fr(comp.get("date_naissance", ""))
        ws.cell(row=row, column=3, value=date_fr).border = thin_border
        ws.cell(row=row, column=4, value=comp.get("sexe", "")).border = thin_border
        ws.cell(row=row, column=5, value=comp.get("poids_declare", "")).border = thin_border
        ws.cell(row=row, column=6, value=comp.get("poids_officiel", "")).border = thin_border
        ws.cell(row=row, column=7, value=comp.get("club", "")).border = thin_border
        ws.cell(row=row, column=8, value=cat_dict.get(comp.get("categorie_id"), "Non assignée")).border = thin_border
        ws.cell(row=row, column=9, value="Oui" if comp.get("pese") else "Non").border = thin_border
        ws.cell(row=row, column=10, value="Oui" if comp.get("surclasse") else "Non").border = thin_border
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 15, 15, 8, 15, 15, 20, 30, 8, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # Sauvegarder dans un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"competiteurs_{competition['nom'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/excel/competiteurs/template")
async def get_import_template(user: User = Depends(get_current_user)):
    """Télécharge le template Excel pour l'import de compétiteurs"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Compétiteurs"
    
    # Style pour l'en-tête
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes requis
    headers = ["Nom*", "Prénom*", "Date de naissance*", "Sexe*", "Poids déclaré*", "Club*", "Surclassé"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Ajouter des exemples avec format français JJ/MM/AAAA
    example_data = [
        ["DUPONT", "Jean", "15/05/2010", "M", 35.5, "Taekwondo Club Paris", "Non"],
        ["MARTIN", "Marie", "20/08/2012", "F", 28.0, "Taekwondo Club Lyon", "Non"],
        ["DURAND", "Pierre", "10/03/2008", "M", 45.0, "Taekwondo Club Marseille", "Oui"],
    ]
    
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    # Validation pour le sexe
    dv_sexe = DataValidation(type="list", formula1='"M,F"', allow_blank=False)
    dv_sexe.error = "Veuillez entrer M ou F"
    dv_sexe.errorTitle = "Sexe invalide"
    ws.add_data_validation(dv_sexe)
    dv_sexe.add("D2:D1000")
    
    # Validation pour surclassé
    dv_surclasse = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    ws.add_data_validation(dv_surclasse)
    dv_surclasse.add("G2:G1000")
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 15, 18, 8, 15, 25, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # Feuille d'instructions
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "INSTRUCTIONS POUR L'IMPORT DES COMPÉTITEURS",
        "",
        "Colonnes obligatoires (marquées *):",
        "- Nom: Nom de famille du compétiteur",
        "- Prénom: Prénom du compétiteur",
        "- Date de naissance: Format JJ/MM/AAAA (ex: 15/05/2010)",
        "- Sexe: M pour Masculin, F pour Féminin",
        "- Poids déclaré: Poids en kg (ex: 35.5)",
        "- Club: Nom du club",
        "",
        "Colonne optionnelle:",
        "- Surclassé: Oui ou Non (défaut: Non)",
        "",
        "Notes:",
        "- Ne modifiez pas la ligne d'en-tête",
        "- La catégorie sera attribuée automatiquement selon l'âge et le poids",
        "- Si Surclassé=Oui, l'admin devra attribuer manuellement la catégorie",
    ]
    for row, text in enumerate(instructions, 1):
        ws2.cell(row=row, column=1, value=text)
    ws2.column_dimensions['A'].width = 60
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import_competiteurs.xlsx"}
    )

@api_router.post("/excel/competiteurs/import/{competition_id}")
async def import_competiteurs_excel(
    competition_id: str,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user)
):
    """Importe des compétiteurs depuis un fichier Excel"""
    if not await user_can_access_competition(user, competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    import pandas as pd
    from openpyxl import load_workbook
    
    # Vérifier que la compétition existe et est active
    competition = await db.competitions.find_one({"competition_id": competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    if competition.get("statut") != "active" and user.role not in ["admin", "master"]:
        raise HTTPException(status_code=400, detail="La compétition n'est plus active")
    
    # Vérifier le type de fichier
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format Excel (.xlsx ou .xls)")
    
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Lire les en-têtes
        headers = [cell.value for cell in ws[1]]
        
        # Mapper les colonnes (ignorer les astérisques)
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                clean_h = h.replace("*", "").strip().lower()
                col_map[clean_h] = i
        
        # Vérifier les colonnes requises
        required = ["nom", "prénom", "date de naissance", "sexe", "poids déclaré", "club"]
        missing = [r for r in required if r not in col_map]
        if missing:
            raise HTTPException(status_code=400, detail=f"Colonnes manquantes: {', '.join(missing)}")
        
        # Importer les compétiteurs
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[col_map["nom"]]:  # Ligne vide
                continue
            
            try:
                nom = str(row[col_map["nom"]]).strip()
                prenom = str(row[col_map["prénom"]]).strip()
                date_naissance = row[col_map["date de naissance"]]
                sexe = str(row[col_map["sexe"]]).strip().upper()
                poids_declare = float(row[col_map["poids déclaré"]])
                club = str(row[col_map["club"]]).strip()
                surclasse = str(row[col_map.get("surclassé", -1)] or "Non").strip().lower() in ["oui", "yes", "true", "1"]
                
                # Convertir la date si nécessaire
                if isinstance(date_naissance, datetime):
                    date_naissance = date_naissance.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_naissance).strip()
                    # Convertir du format français JJ/MM/AAAA au format ISO
                    date_naissance = date_fr_to_iso(date_str)
                
                # Valider le sexe
                if sexe not in ["M", "F"]:
                    errors.append(f"Ligne {row_idx}: Sexe invalide '{sexe}' (doit être M ou F)")
                    continue
                
                # Valider la date
                try:
                    datetime.strptime(date_naissance, "%Y-%m-%d")
                except ValueError:
                    errors.append(f"Ligne {row_idx}: Date de naissance invalide '{date_naissance}' (format: JJ/MM/AAAA)")
                    continue
                
                # Créer le compétiteur
                comp = Competiteur(
                    competition_id=competition_id,
                    nom=nom,
                    prenom=prenom,
                    date_naissance=date_naissance,
                    sexe=sexe,
                    poids_declare=poids_declare,
                    club=club,
                    surclasse=surclasse,
                    created_by=user.user_id
                )
                comp_dict = comp.model_dump()
                comp_dict["created_at"] = comp_dict["created_at"].isoformat()
                
                # Attribution automatique de la catégorie
                if not surclasse:
                    categorie_id = await assign_categorie(comp_dict, competition_id)
                    comp_dict["categorie_id"] = categorie_id
                
                await db.competiteurs.insert_one(comp_dict)
                imported += 1
                
            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")
        
        return {
            "message": f"{imported} compétiteur(s) importé(s)",
            "imported": imported,
            "errors": errors[:10] if errors else [],  # Limiter à 10 erreurs
            "total_errors": len(errors)
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier: {str(e)}")

# ============ COMBATS MANUELS ENDPOINTS ============

async def check_combat_pret(combat_id: str) -> bool:
    """
    Vérifie si un combat est prêt à être lancé.
    Un combat est prêt si:
    - Les deux combattants sont définis (rouge_id et bleu_id)
    - Les combattants ne sont pas disqualifiés
    - Le combat n'est pas déjà terminé ou en cours
    """
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        return False
    
    if combat.get("termine") or combat.get("statut") in ["en_cours", "termine"]:
        return False
    
    rouge_id = combat.get("rouge_id")
    bleu_id = combat.get("bleu_id")
    
    if not rouge_id or not bleu_id:
        return False
    
    # Vérifier que les compétiteurs existent et ne sont pas disqualifiés
    rouge = await db.competiteurs.find_one({"competiteur_id": rouge_id}, {"_id": 0})
    bleu = await db.competiteurs.find_one({"competiteur_id": bleu_id}, {"_id": 0})
    
    if not rouge or not bleu:
        return False
    
    if rouge.get("disqualifie") or bleu.get("disqualifie"):
        return False
    
    return True

async def update_combat_pret_status(combat_id: str):
    """Met à jour le statut 'pret' d'un combat"""
    pret = await check_combat_pret(combat_id)
    await db.combats.update_one(
        {"combat_id": combat_id},
        {"$set": {"pret": pret}}
    )
    return pret

async def propager_vainqueur_manuel(combat: dict, vainqueur_id: str):
    """
    Propage le vainqueur vers le combat suivant (version pour combats manuels).
    Utilise les champs combat_suivant_id et combat_suivant_slot.
    """
    combat_suivant_id = combat.get("combat_suivant_id")
    combat_suivant_slot = combat.get("combat_suivant_slot")
    
    if combat_suivant_id and combat_suivant_slot:
        field = "rouge_id" if combat_suivant_slot == "rouge" else "bleu_id"
        await db.combats.update_one(
            {"combat_id": combat_suivant_id},
            {"$set": {field: vainqueur_id}}
        )
        # Mettre à jour le statut prêt du combat suivant
        await update_combat_pret_status(combat_suivant_id)

@api_router.post("/combats-manuels")
async def create_combat_manuel(data: CombatManuelCreate, user: User = Depends(require_admin)):
    """
    Crée un nouveau combat en mode manuel.
    Permet de définir directement les combattants ou de laisser vide pour connexion ultérieure.
    """
    # Vérifier que la compétition existe
    competition = await db.competitions.find_one({"competition_id": data.competition_id}, {"_id": 0})
    if not competition:
        raise HTTPException(status_code=404, detail="Compétition non trouvée")
    
    # Vérifier que la catégorie existe
    categorie = await db.categories.find_one({"categorie_id": data.categorie_id}, {"_id": 0})
    if not categorie:
        raise HTTPException(status_code=404, detail="Catégorie non trouvée")
    
    # Vérifier les compétiteurs si fournis
    if data.rouge_id:
        rouge = await db.competiteurs.find_one({"competiteur_id": data.rouge_id}, {"_id": 0})
        if not rouge:
            raise HTTPException(status_code=404, detail="Compétiteur rouge non trouvé")
    
    if data.bleu_id:
        bleu = await db.competiteurs.find_one({"competiteur_id": data.bleu_id}, {"_id": 0})
        if not bleu:
            raise HTTPException(status_code=404, detail="Compétiteur bleu non trouvé")
    
    # Calculer l'ordre maximum actuel pour cette catégorie
    max_ordre = await db.combats.find_one(
        {"categorie_id": data.categorie_id},
        {"_id": 0, "ordre": 1},
        sort=[("ordre", -1)]
    )
    nouvel_ordre = (max_ordre.get("ordre", 0) if max_ordre else 0) + 1
    
    # Créer le combat
    combat = Combat(
        competition_id=data.competition_id,
        categorie_id=data.categorie_id,
        tour=data.tour,
        position=data.position,
        ordre=nouvel_ordre,
        rouge_id=data.rouge_id,
        bleu_id=data.bleu_id,
        nom_personnalise=data.nom_personnalise,
        aire_id=data.aire_id,
        mode_creation="manuel",
        pret=bool(data.rouge_id and data.bleu_id)
    )
    
    combat_dict = combat.model_dump()
    combat_dict["created_at"] = combat_dict["created_at"].isoformat()
    
    await db.combats.insert_one(combat_dict)
    combat_dict.pop("_id", None)
    
    return combat_dict

@api_router.put("/combats-manuels/{combat_id}")
async def update_combat_manuel(combat_id: str, data: CombatManuelUpdate, user: User = Depends(require_admin)):
    """
    Met à jour un combat manuel.
    Permet de modifier les combattants, le tour, la position, etc.
    """
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    if combat.get("termine"):
        raise HTTPException(status_code=400, detail="Impossible de modifier un combat terminé")
    
    # Préparer les données de mise à jour
    update_data = {}
    
    if data.tour is not None:
        update_data["tour"] = data.tour
    if data.position is not None:
        update_data["position"] = data.position
    if data.nom_personnalise is not None:
        update_data["nom_personnalise"] = data.nom_personnalise
    if data.aire_id is not None:
        update_data["aire_id"] = data.aire_id
    if data.ordre is not None:
        update_data["ordre"] = data.ordre
    
    # Gestion des combattants
    if data.rouge_id is not None:
        if data.rouge_id == "":
            update_data["rouge_id"] = None
        else:
            rouge = await db.competiteurs.find_one({"competiteur_id": data.rouge_id}, {"_id": 0})
            if not rouge:
                raise HTTPException(status_code=404, detail="Compétiteur rouge non trouvé")
            update_data["rouge_id"] = data.rouge_id
    
    if data.bleu_id is not None:
        if data.bleu_id == "":
            update_data["bleu_id"] = None
        else:
            bleu = await db.competiteurs.find_one({"competiteur_id": data.bleu_id}, {"_id": 0})
            if not bleu:
                raise HTTPException(status_code=404, detail="Compétiteur bleu non trouvé")
            update_data["bleu_id"] = data.bleu_id
    
    if update_data:
        await db.combats.update_one(
            {"combat_id": combat_id},
            {"$set": update_data}
        )
    
    # Mettre à jour le statut prêt
    await update_combat_pret_status(combat_id)
    
    updated = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    return updated

@api_router.delete("/combats-manuels/{combat_id}")
async def delete_combat_manuel(combat_id: str, user: User = Depends(require_admin)):
    """
    Supprime un combat manuel.
    Déconnecte automatiquement les combats liés.
    """
    combat = await db.combats.find_one({"combat_id": combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    if combat.get("termine"):
        raise HTTPException(status_code=400, detail="Impossible de supprimer un combat terminé")
    
    # Si ce combat est une source pour un autre, déconnecter
    await db.combats.update_many(
        {"combat_source_rouge_id": combat_id},
        {"$set": {"combat_source_rouge_id": None, "rouge_id": None}}
    )
    await db.combats.update_many(
        {"combat_source_bleu_id": combat_id},
        {"$set": {"combat_source_bleu_id": None, "bleu_id": None}}
    )
    
    # Déconnecter du combat suivant si lié
    if combat.get("combat_suivant_id"):
        await db.combats.update_one(
            {"combat_id": combat.get("combat_suivant_id")},
            {"$set": {
                f"combat_source_{combat.get('combat_suivant_slot')}_id": None
            }}
        )
    
    # Supprimer le combat
    await db.combats.delete_one({"combat_id": combat_id})
    
    # Mettre à jour les statuts prêts des combats affectés
    combats_affectes = await db.combats.find(
        {"$or": [
            {"combat_source_rouge_id": combat_id},
            {"combat_source_bleu_id": combat_id}
        ]},
        {"_id": 0, "combat_id": 1}
    ).to_list(100)
    
    for c in combats_affectes:
        await update_combat_pret_status(c["combat_id"])
    
    return {"message": "Combat supprimé", "combat_id": combat_id}

@api_router.post("/combats-manuels/connecter")
async def connecter_combats(data: ConnexionCombatRequest, user: User = Depends(require_admin)):
    """
    Connecte le vainqueur d'un combat source vers un slot d'un combat destination.
    
    Exemple: Après combat A, le vainqueur va en position "rouge" du combat B.
    """
    if data.slot_destination not in ["rouge", "bleu"]:
        raise HTTPException(status_code=400, detail="Le slot doit être 'rouge' ou 'bleu'")
    
    # Vérifier que les combats existent
    combat_source = await db.combats.find_one({"combat_id": data.combat_source_id}, {"_id": 0})
    if not combat_source:
        raise HTTPException(status_code=404, detail="Combat source non trouvé")
    
    combat_destination = await db.combats.find_one({"combat_id": data.combat_destination_id}, {"_id": 0})
    if not combat_destination:
        raise HTTPException(status_code=404, detail="Combat destination non trouvé")
    
    # Vérifier qu'ils appartiennent à la même compétition
    if combat_source["competition_id"] != combat_destination["competition_id"]:
        raise HTTPException(status_code=400, detail="Les combats doivent appartenir à la même compétition")
    
    # Vérifier que le slot n'est pas déjà occupé par une connexion
    source_field = f"combat_source_{data.slot_destination}_id"
    if combat_destination.get(source_field):
        raise HTTPException(
            status_code=400, 
            detail=f"Le slot {data.slot_destination} est déjà connecté à un autre combat"
        )
    
    # Mettre à jour le combat source avec la destination
    await db.combats.update_one(
        {"combat_id": data.combat_source_id},
        {"$set": {
            "combat_suivant_id": data.combat_destination_id,
            "combat_suivant_slot": data.slot_destination
        }}
    )
    
    # Mettre à jour le combat destination avec la source
    await db.combats.update_one(
        {"combat_id": data.combat_destination_id},
        {"$set": {source_field: data.combat_source_id}}
    )
    
    # Si le combat source est déjà terminé, propager le vainqueur
    if combat_source.get("termine") and combat_source.get("vainqueur_id"):
        field = "rouge_id" if data.slot_destination == "rouge" else "bleu_id"
        await db.combats.update_one(
            {"combat_id": data.combat_destination_id},
            {"$set": {field: combat_source["vainqueur_id"]}}
        )
        await update_combat_pret_status(data.combat_destination_id)
    
    return {
        "message": "Combats connectés",
        "source": data.combat_source_id,
        "destination": data.combat_destination_id,
        "slot": data.slot_destination
    }

@api_router.post("/combats-manuels/deconnecter")
async def deconnecter_combat(data: DeconnexionCombatRequest, user: User = Depends(require_admin)):
    """
    Déconnecte un combat source d'un slot de destination.
    """
    if data.slot not in ["rouge", "bleu"]:
        raise HTTPException(status_code=400, detail="Le slot doit être 'rouge' ou 'bleu'")
    
    combat_destination = await db.combats.find_one({"combat_id": data.combat_destination_id}, {"_id": 0})
    if not combat_destination:
        raise HTTPException(status_code=404, detail="Combat destination non trouvé")
    
    source_field = f"combat_source_{data.slot}_id"
    combat_source_id = combat_destination.get(source_field)
    
    if not combat_source_id:
        raise HTTPException(status_code=400, detail=f"Aucun combat connecté au slot {data.slot}")
    
    # Déconnecter le combat source
    await db.combats.update_one(
        {"combat_id": combat_source_id},
        {"$set": {"combat_suivant_id": None, "combat_suivant_slot": None}}
    )
    
    # Nettoyer le combat destination
    competiteur_field = "rouge_id" if data.slot == "rouge" else "bleu_id"
    await db.combats.update_one(
        {"combat_id": data.combat_destination_id},
        {"$set": {
            source_field: None,
            competiteur_field: None
        }}
    )
    
    await update_combat_pret_status(data.combat_destination_id)
    
    return {
        "message": "Connexion supprimée",
        "destination": data.combat_destination_id,
        "slot": data.slot
    }

@api_router.get("/combats-manuels/prets/{competition_id}")
async def get_combats_prets(competition_id: str, user: User = Depends(get_current_user)):
    """
    Liste tous les combats prêts à être lancés pour une compétition.
    Un combat est prêt si les deux combattants sont définis et non disqualifiés.
    """
    if not await user_can_access_competition(user, competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    # Mettre à jour le statut prêt de tous les combats non terminés
    combats_a_verifier = await db.combats.find(
        {"competition_id": competition_id, "termine": False},
        {"_id": 0, "combat_id": 1}
    ).to_list(1000)
    
    for c in combats_a_verifier:
        await update_combat_pret_status(c["combat_id"])
    
    # Récupérer les combats prêts
    combats = await db.combats.find(
        {
            "competition_id": competition_id,
            "pret": True,
            "termine": False,
            "statut": "a_venir"
        },
        {"_id": 0}
    ).sort([("ordre", 1)]).to_list(500)
    
    # ============ OPTIMISATION: Batch fetch pour éviter N+1 queries ============
    competiteur_ids = set()
    categorie_ids = set()
    aire_ids = set()
    for combat in combats:
        if combat.get("rouge_id"):
            competiteur_ids.add(combat["rouge_id"])
        if combat.get("bleu_id"):
            competiteur_ids.add(combat["bleu_id"])
        if combat.get("categorie_id"):
            categorie_ids.add(combat["categorie_id"])
        if combat.get("aire_id"):
            aire_ids.add(combat["aire_id"])
    
    competiteurs_list = await db.competiteurs.find(
        {"competiteur_id": {"$in": list(competiteur_ids)}},
        {"_id": 0, "competiteur_id": 1, "nom": 1, "prenom": 1, "club": 1}
    ).to_list(1000)
    competiteurs_map = {c["competiteur_id"]: c for c in competiteurs_list}
    
    categories_list = await db.categories.find(
        {"categorie_id": {"$in": list(categorie_ids)}},
        {"_id": 0, "categorie_id": 1, "nom": 1}
    ).to_list(200)
    categories_map = {c["categorie_id"]: c for c in categories_list}
    
    aires_list = await db.aires_combat.find(
        {"aire_id": {"$in": list(aire_ids)}},
        {"_id": 0, "aire_id": 1, "nom": 1, "numero": 1}
    ).to_list(50)
    aires_map = {a["aire_id"]: a for a in aires_list}
    
    # Enrichir avec les infos (lookup rapide)
    for combat in combats:
        if combat.get("rouge_id"):
            combat["rouge"] = competiteurs_map.get(combat["rouge_id"])
        
        if combat.get("bleu_id"):
            combat["bleu"] = competiteurs_map.get(combat["bleu_id"])
        
        categorie = categories_map.get(combat.get("categorie_id"))
        combat["categorie_nom"] = categorie["nom"] if categorie else "Inconnue"
        
        if combat.get("aire_id"):
            combat["aire"] = aires_map.get(combat["aire_id"])
    
    return combats

@api_router.get("/combats-manuels/non-assignes/{competition_id}")
async def get_combats_non_assignes(competition_id: str, user: User = Depends(get_current_user)):
    """
    Liste les combats prêts mais non assignés à une aire de combat.
    """
    if not await user_can_access_competition(user, competition_id):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    combats = await db.combats.find(
        {
            "competition_id": competition_id,
            "pret": True,
            "termine": False,
            "statut": "a_venir",
            "$or": [{"aire_id": None}, {"aire_id": ""}]
        },
        {"_id": 0}
    ).sort([("ordre", 1)]).to_list(500)
    
    # Enrichir
    for combat in combats:
        if combat.get("rouge_id"):
            rouge = await db.competiteurs.find_one(
                {"competiteur_id": combat["rouge_id"]},
                {"_id": 0, "nom": 1, "prenom": 1, "club": 1}
            )
            combat["rouge"] = rouge
        
        if combat.get("bleu_id"):
            bleu = await db.competiteurs.find_one(
                {"competiteur_id": combat["bleu_id"]},
                {"_id": 0, "nom": 1, "prenom": 1, "club": 1}
            )
            combat["bleu"] = bleu
        
        categorie = await db.categories.find_one(
            {"categorie_id": combat["categorie_id"]},
            {"_id": 0, "nom": 1}
        )
        combat["categorie_nom"] = categorie["nom"] if categorie else "Inconnue"
    
    return combats

@api_router.post("/combats-manuels/assigner-aire")
async def assigner_combat_aire(data: AssignerAireRequest, user: User = Depends(require_admin)):
    """
    Assigne un combat à une aire de combat spécifique.
    """
    combat = await db.combats.find_one({"combat_id": data.combat_id}, {"_id": 0})
    if not combat:
        raise HTTPException(status_code=404, detail="Combat non trouvé")
    
    aire = await db.aires_combat.find_one({"aire_id": data.aire_id}, {"_id": 0})
    if not aire:
        raise HTTPException(status_code=404, detail="Aire de combat non trouvée")
    
    # Vérifier que l'aire est active
    if aire.get("statut") not in ["active", None]:
        raise HTTPException(status_code=400, detail="Cette aire de combat n'est pas active")
    
    # Calculer l'ordre dans l'aire
    max_ordre_aire = await db.combats.find_one(
        {"aire_id": data.aire_id, "termine": False},
        {"_id": 0, "ordre": 1},
        sort=[("ordre", -1)]
    )
    nouvel_ordre = (max_ordre_aire.get("ordre", 0) if max_ordre_aire else 0) + 1
    
    await db.combats.update_one(
        {"combat_id": data.combat_id},
        {"$set": {"aire_id": data.aire_id, "ordre": nouvel_ordre}}
    )
    
    updated = await db.combats.find_one({"combat_id": data.combat_id}, {"_id": 0})
    return updated

@api_router.post("/combats-manuels/distribution-auto/{competition_id}")
async def distribution_auto_combats(competition_id: str, user: User = Depends(require_admin)):
    """
    Distribue automatiquement les combats prêts et non assignés sur les aires disponibles.
    Utilise un algorithme round-robin pour équilibrer la charge.
    """
    # Récupérer les aires actives
    aires = await db.aires_combat.find(
        {"competition_id": competition_id, "statut": {"$in": ["active", None]}},
        {"_id": 0}
    ).sort("numero", 1).to_list(20)
    
    if not aires:
        raise HTTPException(status_code=400, detail="Aucune aire de combat active")
    
    # Récupérer les combats prêts non assignés
    combats = await db.combats.find(
        {
            "competition_id": competition_id,
            "pret": True,
            "termine": False,
            "statut": "a_venir",
            "$or": [{"aire_id": None}, {"aire_id": ""}]
        },
        {"_id": 0}
    ).sort([("est_finale", 1), ("ordre", 1)]).to_list(500)
    
    if not combats:
        return {"message": "Aucun combat à distribuer", "distribues": 0}
    
    # Distribution round-robin
    nb_aires = len(aires)
    distribues = 0
    
    for i, combat in enumerate(combats):
        aire = aires[i % nb_aires]
        
        # Calculer l'ordre dans l'aire
        max_ordre = await db.combats.find_one(
            {"aire_id": aire["aire_id"], "termine": False},
            {"_id": 0, "ordre": 1},
            sort=[("ordre", -1)]
        )
        nouvel_ordre = (max_ordre.get("ordre", 0) if max_ordre else 0) + 1
        
        await db.combats.update_one(
            {"combat_id": combat["combat_id"]},
            {"$set": {"aire_id": aire["aire_id"], "ordre": nouvel_ordre}}
        )
        distribues += 1
    
    return {
        "message": f"{distribues} combat(s) distribué(s) sur {nb_aires} aire(s)",
        "distribues": distribues,
        "aires": nb_aires
    }

@api_router.get("/combats-manuels/arbre/{categorie_id}")
async def get_arbre_manuel(categorie_id: str, user: User = Depends(get_current_user)):
    """
    Récupère l'arbre de combats d'une catégorie avec les connexions pour affichage graphique.
    Inclut les informations de connexion entre combats pour le drag & drop.
    """
    combats = await db.combats.find({"categorie_id": categorie_id}, {"_id": 0}).to_list(200)
    
    # Enrichir chaque combat
    for combat in combats:
        if combat.get("rouge_id"):
            rouge = await db.competiteurs.find_one(
                {"competiteur_id": combat["rouge_id"]},
                {"_id": 0, "nom": 1, "prenom": 1, "club": 1}
            )
            combat["rouge"] = rouge
        else:
            combat["rouge"] = None
        
        if combat.get("bleu_id"):
            bleu = await db.competiteurs.find_one(
                {"competiteur_id": combat["bleu_id"]},
                {"_id": 0, "nom": 1, "prenom": 1, "club": 1}
            )
            combat["bleu"] = bleu
        else:
            combat["bleu"] = None
        
        if combat.get("vainqueur_id"):
            vainqueur = await db.competiteurs.find_one(
                {"competiteur_id": combat["vainqueur_id"]},
                {"_id": 0, "nom": 1, "prenom": 1}
            )
            combat["vainqueur"] = vainqueur
    
    # Organiser par tour
    arbre = {}
    for combat in combats:
        tour = combat.get("tour", "autre")
        if tour not in arbre:
            arbre[tour] = []
        arbre[tour].append(combat)
    
    # Trier par position dans chaque tour
    for tour in arbre:
        arbre[tour] = sorted(arbre[tour], key=lambda x: x.get("position", 0))
    
    # Créer la liste des connexions
    connexions = []
    for combat in combats:
        if combat.get("combat_suivant_id"):
            connexions.append({
                "source_id": combat["combat_id"],
                "destination_id": combat["combat_suivant_id"],
                "slot": combat.get("combat_suivant_slot")
            })
    
    categorie = await db.categories.find_one({"categorie_id": categorie_id}, {"_id": 0})
    
    return {
        "categorie": categorie,
        "arbre": arbre,
        "connexions": connexions,
        "total_combats": len(combats),
        "combats_termines": len([c for c in combats if c.get("termine")]),
        "combats_prets": len([c for c in combats if c.get("pret") and not c.get("termine")])
    }

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
